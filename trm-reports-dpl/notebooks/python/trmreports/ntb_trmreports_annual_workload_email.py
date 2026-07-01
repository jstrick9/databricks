# Databricks notebook source
# MAGIC %md
# MAGIC # Combined Excel Email Script
# MAGIC
# MAGIC This notebook combines seven annual workload tables into a single Excel workbook with separate tabs:
# MAGIC - **Table 16**: Trademark Applications Filed for Registration and Renewal
# MAGIC - **Table 17**: Summary of Pending Trademark Applications
# MAGIC - **Table 18**: Trademarks Registered, Renewed, and Published Under Section 12(C)
# MAGIC - **Table 19**: Trademark Applications Filed by Residents of the United States
# MAGIC - **Table 20**: Trademarks Registered to Residents of the United States
# MAGIC - **Table 21**: Trademark Applications Filed by Residents of Foreign Countries
# MAGIC - **Table 22**: Trademarks Registered to Residents of Foreign Countries
# MAGIC
# MAGIC All tables are sent as separate tabs in a single Excel file via email.

# COMMAND ----------

# MAGIC %md
# MAGIC ## Config

# COMMAND ----------

# DBTITLE 1,Parameters and Configs
dbutils.widgets.text("dbx_env", "dev")
dbx_env = dbutils.widgets.get("dbx_env").rstrip()

# Add current_fy widget for testing with different fiscal years
import datetime
dbutils.widgets.text("current_fy", str(datetime.datetime.now().year))
current_fy = int(dbutils.widgets.get("current_fy"))

config_file_name = "trmreports-conf.yaml"
config_file = "../../config/" + dbutils.widgets.get("dbx_env") + "/" + config_file_name

print(f"{config_file=},{dbx_env=},{current_fy=}")

# COMMAND ----------

# DBTITLE 1,Execute common function ntbk
# MAGIC %run ./../shared/ntb_common_func_and_params

# COMMAND ----------

# DBTITLE 1,Parameter Values
common_configs = read_yaml(config_file)
reporting_catalog = common_configs["schema"]["reporting_catalog"]
trgt_catalog = common_configs["schema"]["trgt_catalog"]
tmprodvty_catalog = common_configs["schema"]["tmprodvty_catalog"]
primary_email, cc_email = common_configs["alerting"]["annual_workload_tables"]["email"], common_configs["alerting"]["annual_workload_tables"]["cc"]
print(reporting_catalog, trgt_catalog, tmprodvty_catalog, primary_email, cc_email)

# COMMAND ----------

# DBTITLE 1,Print values
print(f"{reporting_catalog=},{trgt_catalog=}")

# COMMAND ----------

# DBTITLE 1,Start Job Control
job_name = "AFR Combined Tables Email Report"
control_dt = begin_job_cntl(f"{reporting_catalog}.silver", job_name, job_start_ts)

# COMMAND ----------

# DBTITLE 1,Install OpenPyxl
# MAGIC %pip install openpyxl

# COMMAND ----------

# DBTITLE 1,Import Libraries
from io import BytesIO
import pandas as pd
import smtplib
import builtins

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from email import encoders
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import COMMASPACE

import datetime

# COMMAND ----------

# MAGIC %md
# MAGIC ## Formatting

# COMMAND ----------

# DBTITLE 1,Excel Formatting: 16, 17, 21, 22
def format_table_16_sheet(ws, df):
    """
    Format a worksheet for Table 16 with specific styling.
    
    Args:
        ws: openpyxl worksheet object
        df: pandas DataFrame with Table 16 data
    """
    # Define styles
    title_font = Font(name="Times New Roman", size=12, bold=True)
    subtitle_font = Font(name="Times New Roman", size=11, bold=True)
    header_font = Font(name="Times New Roman", size=10, bold=True)
    body_font = Font(name="Times New Roman", size=10)
    small_font = Font(name="Times New Roman", size=8)
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Row 1: "TABLE 16"
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "TABLE 16"
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="bottom")
    ws.row_dimensions[1].height = 15

    # Row 2
    ws.merge_cells("A2:D2")
    cell = ws["A2"]
    cell.value = "TRADEMARK APPLICATIONS FILED FOR REGISTRATION AND RENEWAL AND TRADEMARK AFFADAVITS FILED"
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Row 3
    first_year = df["Year"].iloc[0]
    last_year = df["Year"].iloc[-1]
    ws.merge_cells("A3:D3")
    cell = ws["A3"]
    cell.value = f"(FY {first_year} - FY {last_year})"
    cell.font = subtitle_font
    cell.alignment = Alignment(horizontal="center", vertical="bottom")
    ws.row_dimensions[3].height = 15

    # Row 4
    cell = ws["B4"]
    cell.value = "For"
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    cell = ws["C4"]
    cell.value = "¹For"
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    cell = ws["D4"]
    cell.value = "Section 8"
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)

    # Write headers (row 5)
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=5, column=col_idx)
        cell.value = col_name
        cell.font = header_font
        cell.alignment = center_align

    # Write data (starting row 6)
    last_year = df["Year"].iloc[-1]
    body_font_bold = Font(name="Times New Roman", size=10, bold=True)
    
    for row_idx, row in enumerate(df.itertuples(index=False), start=6):
        is_current_year = (row[0] == last_year)
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = body_font_bold if is_current_year else body_font
            cell.alignment = center_align
            if col_idx > 1 and isinstance(value, (int, float)):
                cell.number_format = '#,##0'

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # Add borders
    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="thick", color="000000")
    
    # Two-block high box border around columns A-D, rows 4-5
    for row in range(4, 6):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            border = Border(
                left=thin if col == 1 else None,
                right=thin if col == 4 else None,
                top=thin if row == 4 else None,
                bottom=thin if row == 5 else None,
            )
            cell.border = border

    # Box border around dataframe values
    data_start_row = 6
    data_end_row = data_start_row + len(df) - 1
    for row in range(data_start_row, data_end_row + 1):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            border = Border(
                left=thin if col == 1 else None,
                right=thin if col == 4 else None,
                top=thin if row == data_start_row else None,
                bottom=thin if row == data_end_row else None,
            )
            cell.border = border

    # Thick outside border
    max_row = len(df) + 7
    for row in range(1, max_row + 1):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            border = Border(
                left=thick if col == 1 else None,
                right=thick if col == 4 else None,
                top=thick if row == 1 else None,
                bottom=thick if row == max_row else None,
            )
            cell.border = Border(
                left=border.left if border.left else cell.border.left,
                right=border.right if border.right else cell.border.right,
                top=border.top if border.top else cell.border.top,
                bottom=border.bottom if border.bottom else cell.border.bottom,
            )
    
    # Footer row
    ws.merge_cells(f"A{max_row}:D{max_row}")
    cell = ws.cell(row=max_row, column=1)
    cell.value = "¹Renewal of registration term changed in November 16, 1989 (FY1990) with the implementation of the Trademark Law Reform Act (Pub. l. No. 100–667)."
    cell.font = small_font
    cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)
    ws.row_dimensions[max_row].height = 30

def format_table_17_sheet(ws, df_table_17):
    """
    Write and format Table 17 (Summary of Pending Trademark Applications)
    onto an existing openpyxl worksheet.

    Args:
        ws          : openpyxl Worksheet object
        df_table_17 : pandas DataFrame with columns
                      ['row_num', 'col1', 'col2', 'col3']
                      built from the sorted output_rows list of tuples.

    Formatting applied:
      - Font: Times New Roman throughout
      - Rows 1-3  : size 12, bold, columns A-C merged, center-justified
      - All others: size 10
      - Bold rows : 1, 2, 3, 4, 6, 8, 10, 12, 16, 18, 23, 25
      - Row 4     : top + bottom border
      - Rows 27-28: top + bottom border
      - Rows 4-28 : single outer encompassing border
    """
    BOLD_ROWS = {1, 2, 3, 4, 6, 8, 10, 12, 16, 18, 23, 25}
    TITLE_ROWS = {1, 2, 3}

    thin = Side(style="thin")

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15

    for entry in df_table_17.itertuples(index=False):
        row_num = entry[0]
        values = list(entry[1:])

        is_title = row_num in TITLE_ROWS
        is_bold = row_num in BOLD_ROWS

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = Font(
                name="Times New Roman",
                bold=is_bold,
                size=12 if is_title else 10,
            )
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0"

    for r in TITLE_ROWS:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        cell = ws.cell(row=r, column=1)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="Times New Roman", bold=True, size=12)

    for col in range(1, 4):
        ws.cell(row=4, column=col).border = Border(top=thin, bottom=thin)

    for r in [27, 28]:
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = Border(top=thin, bottom=thin)

    first_row, last_row = 4, 28
    first_col, last_col = 1, 3

    for r in range(first_row, last_row + 1):
        for c in range(first_col, last_col + 1):
            existing = ws.cell(row=r, column=c).border
            top_side = existing.top or None
            bottom_side = existing.bottom or None
            left_side = existing.left or None
            right_side = existing.right or None

            if r == first_row:
                top_side = thin
            if r == last_row:
                bottom_side = thin
            if c == first_col:
                left_side = thin
            if c == last_col:
                right_side = thin

            ws.cell(row=r, column=c).border = Border(
                top=top_side,
                bottom=bottom_side,
                left=left_side,
                right=right_side,
            )

def format_foreign_country_sheet(ws, df, sheet_name, last_5_fys):
    """
    Format a worksheet for Table 21 or Table 22 with specific styling.
    
    Args:
        ws: openpyxl worksheet object
        df: pandas DataFrame with table data
        sheet_name: name of the sheet to determine which table config to use
        last_5_fys: list of fiscal years for subtitle
    """
    # Define styles
    times_font = Font(name="Times New Roman", size=10)
    times_bold = Font(name="Times New Roman", size=10, bold=True)
    header_font = Font(name="Times New Roman", size=12, bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    left_wrap_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    medium_side = Side(style="medium", color="000000")
    number_format = '#,##0'

    ncols = df.shape[1]
    nrows = df.shape[0]

    # Determine which table this is
    if "22" in sheet_name:
        title_text = "TABLE 22"
        subtitle_text = "TRADEMARKS REGISTERED TO RESIDENTS OF FOREIGN COUNTRIES"
    else:
        title_text = "TABLE 21"
        subtitle_text = "TRADEMARK APPLICATIONS FILED BY RESIDENTS OF FOREIGN COUNTRIES AND TERRITORIES"

    fy_text = f"(FY {builtins.min(last_5_fys)} - FY {builtins.max(last_5_fys)})" if last_5_fys else ""
    after_table_text_1 = " - Represents zero."
    after_table_text_2 = "¹Country of Origin Information not available or not indicated in the database (includes African Regional Intellectual Property Organization filings)."

    # Layout constants - Table 22 starts at row 2, Table 21 starts at row 1
    row_offset = 1 if "22" in sheet_name else 0
    title_row = 1 + row_offset
    subtitle_row = 2 + row_offset
    fy_row = 3 + row_offset
    table_start_row = 4 + row_offset
    after_table_row_1 = table_start_row + nrows + 1
    after_table_row_2 = table_start_row + nrows + 2

    # Set column widths
    ws.column_dimensions['A'].width = 27.33
    ws.column_dimensions['B'].width = 9.5
    ws.column_dimensions['C'].width = 9.5
    ws.column_dimensions['D'].width = 9.5
    ws.column_dimensions['E'].width = 9.5
    ws.column_dimensions['F'].width = 9.5

    # Identify numeric columns
    numeric_cols = [
        i for i, col in enumerate(df.columns)
        if pd.api.types.is_numeric_dtype(df[col]) or (
            df[col].dtype == object and
            all(
                (v == "" or v is None or isinstance(v, (int, float)) or 
                 (isinstance(v, str) and v.replace(",", "").replace("-", "").isdigit()))
                for v in df[col]
            )
        )
    ]

    # Title row - merge only first 3 columns
    ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=3)
    cell = ws.cell(row=title_row, column=1)
    cell.value = title_text
    cell.font = header_font
    cell.alignment = center_align

    # Subtitle row - merge only first 3 columns
    ws.row_dimensions[subtitle_row].height = 30
    ws.merge_cells(start_row=subtitle_row, start_column=1, end_row=subtitle_row, end_column=3)
    cell = ws.cell(row=subtitle_row, column=1)
    cell.value = subtitle_text
    cell.font = header_font
    cell.alignment = wrap_align

    # FY row - merge only first 3 columns
    ws.merge_cells(start_row=fy_row, start_column=1, end_row=fy_row, end_column=3)
    cell = ws.cell(row=fy_row, column=1)
    cell.value = fy_text
    cell.font = header_font
    cell.alignment = center_align

    # Write header row
    start_row = table_start_row
    start_col = 1

    for c_idx, col_name in enumerate(df.columns):
        header_value = "\u200B" + str(col_name).replace("FY", "") if str(col_name).startswith("FY") else str(col_name)
        cell = ws.cell(row=start_row, column=start_col + c_idx)
        cell.value = header_value
        cell.font = times_bold
        cell.alignment = left_align if c_idx == 0 else center_align

    # Header border
    for c_idx in range(start_col, start_col + ncols):
        cell = ws.cell(row=start_row, column=c_idx)
        is_left = (c_idx == start_col)
        is_right = (c_idx == start_col + ncols - 1)
        cell.border = Border(
            top=medium_side,
            bottom=medium_side,
            left=medium_side if is_left else None,
            right=medium_side if is_right else None,
        )

    # Write data rows
    for r_idx, row_data in enumerate(df.values):
        excel_row = start_row + 1 + r_idx
        for c_idx, cell_value in enumerate(row_data):
            cell_value_to_write = "-" if cell_value == 0 else cell_value
            cell = ws.cell(row=excel_row, column=start_col + c_idx)
            cell.value = cell_value_to_write
            cell.font = times_bold if excel_row == start_row + 2 else times_font
            cell.alignment = left_align if c_idx == 0 else center_align

            if c_idx in numeric_cols and cell_value not in ("", None, 0):
                cell.number_format = number_format

    # Border around data area
    data_start_row = start_row + 1
    data_end_row = start_row + nrows
    data_start_col = start_col
    data_end_col = start_col + ncols - 1

    for r in range(data_start_row, data_end_row + 1):
        for c in range(data_start_col, data_end_col + 1):
            cell = ws.cell(row=r, column=c)
            is_top = (r == data_start_row)
            is_bottom = (r == data_end_row)
            is_left = (c == data_start_col)
            is_right = (c == data_end_col)

            if is_top or is_bottom or is_left or is_right:
                cell.border = Border(
                    top=medium_side if is_top else None,
                    bottom=medium_side if is_bottom else None,
                    left=medium_side if is_left else None,
                    right=medium_side if is_right else None,
                )

    # Notes below table
    ws.merge_cells(start_row=after_table_row_1, start_column=1, end_row=after_table_row_1, end_column=ncols)
    cell = ws.cell(row=after_table_row_1, column=1)
    cell.value = after_table_text_1
    cell.font = times_font
    cell.alignment = left_wrap_align

    ws.merge_cells(start_row=after_table_row_2, start_column=1, end_row=after_table_row_2, end_column=ncols)
    cell = ws.cell(row=after_table_row_2, column=1)
    cell.value = after_table_text_2
    cell.font = times_font
    cell.alignment = left_wrap_align

    ws.row_dimensions[after_table_row_1].height = 18
    ws.row_dimensions[after_table_row_2].height = 36

    # Border around notes
    for r in [after_table_row_1, after_table_row_2]:
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            is_top = (r == after_table_row_1)
            is_bottom = (r == after_table_row_2)
            is_left = (c == 1)
            is_right = (c == ncols)

            if is_top or is_bottom or is_left or is_right:
                cell.border = Border(
                    top=medium_side if is_top else None,
                    bottom=medium_side if is_bottom else None,
                    left=medium_side if is_left else None,
                    right=medium_side if is_right else None,
                )

# COMMAND ----------

# DBTITLE 1,Excel Formatting: 18, 19, 20
def format_table_18_sheet(ws, df):
    """
    Format a worksheet for Table 18 with specific styling requirements.
    
    Args:
        ws: openpyxl worksheet object
        df: pandas DataFrame with Table 18 data
    """
    from openpyxl.styles.fonts import Font as FontClass
    
    # Define styles
    title_font = Font(name="Times New Roman", size=12, bold=True)
    subtitle_font = Font(name="Times New Roman", size=11, bold=True)
    header_font = Font(name="Times New Roman", size=10, bold=True)
    body_font = Font(name="Times New Roman", size=10)
    footer_font = Font(name="Times New Roman", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Row 1: "TABLE 18" (single height, no forced height)
    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "TABLE 18"
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="bottom")

    # Row 2: Title with superscript
    ws.merge_cells("A2:D2")
    cell = ws["A2"]
    cell.value = "TRADEMARKS REGISTERED, RENEWED, AND PUBLISHED UNDER SECTION 12(C)¹"
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Row 3: FY range (single height, no forced height)
    first_year = int(df["Year"].iloc[0])
    last_year = int(df["Year"].iloc[-1])
    ws.merge_cells("A3:D3")
    cell = ws["A3"]
    cell.value = f"(FY {first_year} - FY {last_year})"
    cell.font = subtitle_font
    cell.alignment = Alignment(horizontal="center", vertical="bottom")
    
    # Add border around title rows (1-3)
    thin = Side(border_style="thin", color="000000")
    for row in range(1, 4):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            is_top = (row == 1)
            is_bottom = (row == 3)
            is_left = (col == 1)
            is_right = (col == 4)
            cell.border = Border(
                left=thin if is_left else None,
                right=thin if is_right else None,
                top=thin if is_top else None,
                bottom=thin if is_bottom else None
            )

    # Set column widths
    ws.column_dimensions["A"].width = 11.11
    ws.column_dimensions["B"].width = 23.44
    ws.column_dimensions["C"].width = 17.44
    ws.column_dimensions["D"].width = 19.44

    # Row 4 headers
    ws.cell(row=4, column=1).value = "Year"
    ws.cell(row=4, column=1).font = header_font
    ws.cell(row=4, column=1).alignment = center_align
    
    ws.cell(row=4, column=2).value = "Certificates of Registration"
    ws.cell(row=4, column=2).font = header_font
    ws.cell(row=4, column=2).alignment = center_align
    
    ws.cell(row=4, column=3).value = "Renewed²"
    ws.cell(row=4, column=3).font = header_font
    ws.cell(row=4, column=3).alignment = center_align
    
    ws.cell(row=4, column=4).value = "Registrations"
    ws.cell(row=4, column=4).font = header_font
    ws.cell(row=4, column=4).alignment = center_align

    # Row 5 sub-headers
    ws.cell(row=5, column=2).value = "Issued"
    ws.cell(row=5, column=2).font = header_font
    ws.cell(row=5, column=2).alignment = center_align
    
    ws.cell(row=5, column=4).value = "(Including Classes)"
    ws.cell(row=5, column=4).font = header_font
    ws.cell(row=5, column=4).alignment = center_align

    # Write data (starting row 6)
    last_year = df["Year"].iloc[-1]
    body_font_bold = Font(name="Times New Roman", size=10, bold=True)
    data_end_row = 6 + len(df) - 1
    
    for row_idx, row in enumerate(df.itertuples(index=False), start=6):
        is_current_year = (row[0] == last_year)
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Fix "Number Stored as Text" warning for Year column
            if col_idx == 1:
                cell.value = int(value)
            else:
                cell.value = value
            
            # Bold the current year row
            cell.font = body_font_bold if is_current_year else body_font
            cell.alignment = center_align
            
            if col_idx > 1 and isinstance(value, (int, float)):
                cell.number_format = '#,##0'
    
    # Add borders
    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="thick", color="000000")
    
    # Border around each header column (rows 4-5)
    for col in range(1, 5):
        for row in [4, 5]:
            cell = ws.cell(row=row, column=col)
            cell.border = Border(
                left=thin,
                right=thin,
                top=thin if row == 4 else None,
                bottom=thin if row == 5 else None
            )
    
    # Border around each data column
    for col in range(1, 5):
        for row in range(6, data_end_row + 1):
            cell = ws.cell(row=row, column=col)
            is_top = (row == 6)
            is_bottom = (row == data_end_row)
            cell.border = Border(
                left=thin,
                right=thin,
                top=thin if is_top else None,
                bottom=thin if is_bottom else None
            )
    
    # Footer box (dynamic position)
    footer_row_1 = data_end_row + 1  # Empty row
    footer_row_2 = data_end_row + 2  # First footnote
    footer_row_3 = data_end_row + 3  # Second footnote
    
    # Footer row 2: "¹Includes withdrawn numbers"
    ws.merge_cells(f"A{footer_row_2}:D{footer_row_2}")
    cell = ws.cell(row=footer_row_2, column=1)
    cell.value = "¹Includes withdrawn numbers"
    cell.font = footer_font
    cell.alignment = left_align
    
    # Footer row 3: Long footnote with superscript
    ws.merge_cells(f"A{footer_row_3}:D{footer_row_3}")
    cell = ws.cell(row=footer_row_3, column=1)
    cell.value = "²Includes renewals that were affected by the reduction of the renewal term of registration from 20 years to 10 years as a result of the implementation in November 16, 1989 (FY 1990) of the Trademark Law Reform Act (Public Law No. 100—667)."
    cell.font = footer_font
    cell.alignment = left_align
    ws.row_dimensions[footer_row_3].height = 60  # Tall enough for wrapped text
    
    # Border around footer box (rows footer_row_1 to footer_row_3)
    for row in range(footer_row_1, footer_row_3 + 1):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            is_top = (row == footer_row_1)
            is_bottom = (row == footer_row_3)
            is_left = (col == 1)
            is_right = (col == 4)
            cell.border = Border(
                left=thin if is_left else None,
                right=thin if is_right else None,
                top=thin if is_top else None,
                bottom=thin if is_bottom else None
            )


def format_state_table_sheet(ws, df, table_number, table_title, current_fy):
    """
    Format a worksheet for Tables 19 or 20 with multi-column state layout.
    
    Args:
        ws: openpyxl worksheet object
        df: pandas DataFrame with state data (State, FY column)
        table_number: "19" or "20"
        table_title: Full title for the table
        current_fy: Current fiscal year
    """
    # Define styles
    title_font = Font(name="Times New Roman", size=12, bold=True)
    header_font = Font(name="Times New Roman", size=10, bold=True)
    body_font = Font(name="Times New Roman", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # Row 1: "TABLE XX" (single line)
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = f"TABLE {table_number}"
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="bottom")

    # Row 2: Title (single line)
    ws.merge_cells("A2:H2")
    cell = ws["A2"]
    cell.value = table_title
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)

    # Row 3: FY Year (single line)
    ws.merge_cells("A3:H3")
    cell = ws["A3"]
    cell.value = f"(FY {current_fy})"
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="bottom")
    
    # Add thin border around title rows (1-3)
    thin = Side(border_style="thin", color="000000")
    for row in range(1, 4):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            is_top = (row == 1)
            is_bottom = (row == 3)
            is_left = (col == 1)
            is_right = (col == 8)
            cell.border = Border(
                left=thin if is_left else None,
                right=thin if is_right else None,
                top=thin if is_top else None,
                bottom=thin if is_bottom else None
            )

    # Column headers (row 4)
    headers = ["State/Territory", current_fy, "", "State/Territory", current_fy, "", "State/Territory", current_fy]
    for col_idx, header in enumerate(headers, start=1):
        if header:  # Skip blank columns
            cell = ws.cell(row=4, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.alignment = center_align if col_idx % 3 == 2 else left_align

    # Add thin border around entire header row (row 4) - single box around A to H
    for col in range(1, 9):
        cell = ws.cell(row=4, column=col)
        is_left = (col == 1)
        is_right = (col == 8)
        cell.border = Border(
            left=thin if is_left else None,
            right=thin if is_right else None,
            top=thin,
            bottom=thin
        )
    
    # Set column widths
    ws.column_dimensions["A"].width = 19.11  # State col 1
    ws.column_dimensions["B"].width = 8.11   # FY col 1
    ws.column_dimensions["C"].width = 1.33   # Blank col
    ws.column_dimensions["D"].width = 19.11  # State col 2
    ws.column_dimensions["E"].width = 8.11   # FY col 2
    ws.column_dimensions["F"].width = 1.33   # Blank col
    ws.column_dimensions["G"].width = 19.11  # State col 3
    ws.column_dimensions["H"].width = 8.11   # FY col 3

    # Get fiscal year column name (should be like "2026" or similar)
    fy_col = [col for col in df.columns if col != "State"][0]
    
    # Calculate total
    total_value = df[fy_col].sum()
    
    # Sort states alphabetically, but put "OTHER¹" at the end
    df_sorted = df.copy()
    df_sorted['sort_key'] = df_sorted['State'].apply(lambda x: 'ZZZZ' if x == 'OTHER¹' else x)
    df_sorted = df_sorted.sort_values('sort_key').drop('sort_key', axis=1).reset_index(drop=True)
    
    # Layout: Column 1 has Total, blank, then 17 states (19 rows)
    #         Column 2 has 19 states
    #         Column 3 has remaining states
    
    current_row = 5
    
    # Column 1: Total
    ws.cell(row=current_row, column=1).value = "Total"
    ws.cell(row=current_row, column=1).font = header_font
    ws.cell(row=current_row, column=1).alignment = left_align
    ws.cell(row=current_row, column=2).value = total_value
    ws.cell(row=current_row, column=2).font = body_font
    ws.cell(row=current_row, column=2).alignment = center_align
    ws.cell(row=current_row, column=2).number_format = '#,##0'
    ws.row_dimensions[current_row].height = 19.8
    current_row += 1
    
    # Blank row
    ws.row_dimensions[current_row].height = 19.8
    current_row += 1
    
    # Column 1: First 17 states
    for i in range(builtins.min(17, len(df_sorted))):
        ws.cell(row=current_row, column=1).value = df_sorted.iloc[i]["State"]
        ws.cell(row=current_row, column=1).font = body_font
        ws.cell(row=current_row, column=1).alignment = left_align
        ws.cell(row=current_row, column=2).value = df_sorted.iloc[i][fy_col]
        ws.cell(row=current_row, column=2).font = body_font
        ws.cell(row=current_row, column=2).alignment = center_align
        ws.cell(row=current_row, column=2).number_format = '#,##0'
        ws.row_dimensions[current_row].height = 19.8
        current_row += 1
    
    # Column 2: Next 19 states (starting from row 5)
    start_idx = 17
    current_row = 5
    for i in range(start_idx, builtins.min(start_idx + 19, len(df_sorted))):
        ws.cell(row=current_row, column=4).value = df_sorted.iloc[i]["State"]
        ws.cell(row=current_row, column=4).font = body_font
        ws.cell(row=current_row, column=4).alignment = left_align
        ws.cell(row=current_row, column=5).value = df_sorted.iloc[i][fy_col]
        ws.cell(row=current_row, column=5).font = body_font
        ws.cell(row=current_row, column=5).alignment = center_align
        ws.cell(row=current_row, column=5).number_format = '#,##0'
        current_row += 1
    
    # Column 3: Remaining states (starting from row 5)
    start_idx = 36
    current_row = 5
    for i in range(start_idx, len(df_sorted)):
        ws.cell(row=current_row, column=7).value = df_sorted.iloc[i]["State"]
        ws.cell(row=current_row, column=7).font = body_font
        ws.cell(row=current_row, column=7).alignment = left_align
        ws.cell(row=current_row, column=8).value = df_sorted.iloc[i][fy_col]
        ws.cell(row=current_row, column=8).font = body_font
        ws.cell(row=current_row, column=8).alignment = center_align
        ws.cell(row=current_row, column=8).number_format = '#,##0'
        current_row += 1
    
    # Determine the maximum row used
    max_data_row = 5 + 19  # Column 1 has 19 rows (Total + blank + 17 states)
    
    # Add thin border around entire dataframe
    for row in range(5, max_data_row):
        for col in range(1, 9):
            if col in [3, 6]:  # Skip blank columns for borders
                continue
            cell = ws.cell(row=row, column=col)
            is_top = (row == 5)
            is_bottom = (row == max_data_row - 1)
            is_left = (col == 1 or col == 4 or col == 7)
            is_right = (col == 2 or col == 5 or col == 8)
            
            # Add border on outer edges of dataframe
            cell.border = Border(
                left=thin if (col == 1 and is_left) else None,
                right=thin if (col == 8 and is_right) else None,
                top=thin if is_top else None,
                bottom=thin if is_bottom else None,
            )
    
    # Add footer box with footnote
    footer_row = max_data_row
    ws.merge_cells(f"A{footer_row}:H{footer_row}")
    footer_cell = ws.cell(row=footer_row, column=1)
    footer_cell.value = "¹Includes Puerto Rico, Army Post Office filings, American Samoa, Guam, U.S. Pacific Islands, and miscellaneous U.S. State / No State Indicated in Database"
    footer_cell.font = Font(name="Times New Roman", size=8)
    footer_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[footer_row].height = 24
    
    # Add border around footer box
    for col in range(1, 9):
        cell = ws.cell(row=footer_row, column=col)
        is_left = (col == 1)
        is_right = (col == 8)
        cell.border = Border(
            left=thin if is_left else None,
            right=thin if is_right else None,
            top=thin,
            bottom=thin
        )

# COMMAND ----------

# MAGIC %md
# MAGIC ## Dataframes

# COMMAND ----------

# DBTITLE 1,df16
# Query from static table annual_workload_table_16_18
# Show last 21 years on a rolling basis (2006-2026)
df_tbl_16 = spark.sql(f"""
SELECT 
    fiscal_year AS Year,
    registration_filed_count AS Registration,
    renewal_filed_count AS Renewal,
    s8_affidavit_filed_count AS Affadavit
FROM `{reporting_catalog}`.gold.annual_workload_table_16_18
WHERE fiscal_year >= year(current_date()) - 20
ORDER BY fiscal_year
""").toPandas()

# COMMAND ----------

# DBTITLE 1,df17
fy = f"(FY {datetime.datetime.now().year + (1 if datetime.datetime.now().month >= 10 else 0)})"

DATA_START: int = 5
DATA_END: int = 27

DEFINED_ROWS: set[int] = set(
    [
        6,
        8,
        10,
        12,
        13,
        14,
        16,
        18,
        19,
        20,
        21,
        23,
        25,
    ]
)
UNDEFINED_ROWS: list[int] = [
    i for i in range(DATA_START, DATA_END) if i not in DEFINED_ROWS
]
SUPERSCRIPTS = {
    1: "\u00B9",
    2: "\u00B2",
}

COLUMNS: list[str] = ["Stage of Processing", "Application Files", "Classes"]
TITLE_ROWS: list[tuple[str]] = [
    (1, "TABLE 17", "", ""),
    (2, "SUMMARY OF PENDING TRADEMARK APPLICATIONS", "", ""),
    (3, f"{fy}", "", ""),
]
HEADER_ROW: tuple[str] = [(4, *COLUMNS)]
BLANK_ROWS = [(i, "", "", "") for i in UNDEFINED_ROWS]
FOOTER_ROWS = [
    (
        DATA_END,
        f"{SUPERSCRIPTS[1]} Includes applications pending before the Trademark Trial and Appeal Board and suspended cases.",
        "",
        "",
    ),
    (
        DATA_END + 1,
        f"{SUPERSCRIPTS[2]} Includes all applications in all phases of publication, issue, and registration.",
        "",
        "",
    ),
]

spark.sql(
    f"""
with deduped as (
  select
    serial_num,
    status_code,
    is_counted_noa,
    is_counted_application,
    latest,
    max(num_active_classes) as num_active_classes
  from
    `{reporting_catalog}`.silver.os34_report_status_detail
  group by
    serial_num,
    status_code,
    is_counted_noa,
    is_counted_application,
    latest
),
under_initial_sub_subset as (
  select
    case
      when status_code in (616,640,641,643,645,644,646,647,648,649,661,663,665,666) then '          Amended, Awaiting Action by Examiner (Initial)'
      when status_code in (638) then '          Awaiting First Action by Examiner'
      when status_code in (630, 631) then 'In Preexamination Processing'
      when status_code in (753,756,757,806,807,808,809,810,811,813,812,814,815,816,817) then '          Amended, Awaiting Action by Examiner (Second)'
      when status_code in (744,745,746,747) then '          Administrative Processing of Statements of Use'
      when status_code in (748) then '          Undergoing Second Examination'
      when status_code in (680,681,686,818,819) then 'In Postexamination Processing{SUPERSCRIPTS[2]}'
      when status_code in (760,763,771,774,794,801,802,650,651,652,653,654) then '     Other Pending Applications {SUPERSCRIPTS[1]}'
    end `Stage of Processing`,
    case
      when status_code in (616,640,641,643,645,644,646,647,648,649,661,663,665,666) then 13
      when status_code in (638) then 14
      when status_code in (630, 631) then 8
      when status_code in (753,756,757,806,807,808,809,810,811,813,812,814,815,816,817) then 21
      when status_code in (744,745,746,747) then 19
      when status_code in (748) then 20
      when status_code in (680,681,686,818,819) then 25
      when status_code in (760,763,771,774,794,801,802,650,651,652,653,654) then 23
    end stage,
    count(distinct serial_num) `Application Files`,
    sum(num_active_classes) `Classes`
  from
    deduped
  where
    is_counted_application = true
    and latest = true
    and status_code in (
      616,630,631,638,640,641,643,644,645,646,647,648,649,650,
      651,652,653,654,661,663,665,666,680,681,686,744,745,746,
      747,748,753,756,757,760,763,771,774,794,801,802,806,
      807,808,809,810,811,812,813,814,815,816,817,818,819
    )
  group by
    all
),
under_initial_subset as (
  select
    case
      when status_code in (616,638,640,641,643,645,644,646,647,648,649,661,663,665,666) then '     Applications Under Initial Examination'
      when status_code in (744,745,746,747,748,753,756,757,806,807,808,809,810,811,813,812,814,815,816,817) then '     Applications Under Second Examination'
    end `Stage of Processing`,
    case
      when status_code in (616,638,640,641,643,645,644,646,647,648,649,661,663,665,666) then 12
      when status_code in (744,745,746,747,748,753,756,757,806,807,808,809,810,811,813,812,814,815,816,817) then 18
    end stage,
    count(distinct serial_num) `Application Files`,
    sum(num_active_classes) `Classes`
  from
    deduped
  where
    is_counted_application = true
    and latest = true
    and status_code in (
      616,638,640,641,643,645,644,
      646,647,648,649,661,663,665,
      666,744,745,746,747,748,753,
      756,757,806,807,808,809,810,
      811,813,812,814,815,816,817
    )
  group by
    all
),
itu as (
  select
    '     Intent-to-Use Applications Pending Use' `Stage of Processing`,
    16 stage,
    count(distinct serial_num) `Application Files`,
    sum(num_active_classes) `Classes`
  from
    deduped
  where
    is_counted_application = true
    and is_counted_noa = true
    and latest = true
),
-- Stage 10 = sum of children: stage 12 + 16 + 18 + 23
under_initial as (
  select
    'Under Examination, Total' `Stage of Processing`,
    10 stage,
    sum(`Application Files`) `Application Files`,
    sum(`Classes`) `Classes`
  from (
    select * from itu
    union all
    select * from under_initial_subset
    union all
    select * from under_initial_sub_subset where stage = 23
  )
),
-- Stage 6 = sum of children: stage 8 + 10 + 25
pending_total as (
  select
    'Pending Applications, Total' `Stage of Processing`,
    6 stage,
    sum(`Application Files`) `Application Files`,
    sum(`Classes`) `Classes`
  from (
    select * from under_initial                                    -- stage 10
    union all
    select * from under_initial_sub_subset where stage = 8        -- stage 8 (In Preexamination Processing)
    union all
    select * from under_initial_sub_subset where stage = 25       -- stage 25 (In Postexamination Processing)
  )
),
output as (
  select * from under_initial
  union all
  select * from under_initial_subset
  union all
  select * from under_initial_sub_subset
  union all
  select * from pending_total
  union all
  select * from itu
)
select
  *
from
  output
"""
).createOrReplaceTempView("vw_staging")

data_rows = [
    (row["stage"], row["Stage of Processing"], row["Application Files"], row["Classes"])
    for row in spark.sql("select * from vw_staging").collect()
]
output_rows = sorted(
    TITLE_ROWS + HEADER_ROW + BLANK_ROWS + data_rows + FOOTER_ROWS, key=lambda t: t[0]
)
df_tbl_17 = pd.DataFrame(output_rows, columns=["stage"] + COLUMNS)

# COMMAND ----------

# DBTITLE 1,df18
# Query from static table annual_workload_table_16_18
# Show last 21 years on a rolling basis (2006-2026)
df_tbl_18 = spark.sql(f"""
SELECT 
    fiscal_year AS `Year`,
    certs_reg_issued_count AS `Certificates of Registration Issued`,
    renewed_count AS `Renewed`,
    reg_inc_class_count AS `Classes`
FROM `{reporting_catalog}`.gold.annual_workload_table_16_18
WHERE fiscal_year >= year(current_date()) - 20
ORDER BY fiscal_year
""").toPandas()

# COMMAND ----------

# DBTITLE 1,df19
# Note: 'OTHER¹' includes Puerto Rico, Army Post Office filings, American Samoa, etc. 'OTHER¹' is specified by the Workload tables and has been tested as valid in the database.
# Guam, U.S. Pacific Islands, and miscellaneous U.S. State / No State Indicated in Database

spark.sql("""
CREATE OR REPLACE TEMP VIEW state_lookup AS
  SELECT * FROM VALUES
    ('AL', 'Alabama'),
    ('AK', 'Alaska'),
    ('AZ', 'Arizona'),
    ('AR', 'Arkansas'),
    ('CA', 'California'),
    ('CO', 'Colorado'),
    ('CT', 'Connecticut'),
    ('DE', 'Delaware'),
    ('FL', 'Florida'),
    ('GA', 'Georgia'),
    ('HI', 'Hawaii'),
    ('ID', 'Idaho'),
    ('IL', 'Illinois'),
    ('IN', 'Indiana'),
    ('IA', 'Iowa'),
    ('KS', 'Kansas'),
    ('KY', 'Kentucky'),
    ('LA', 'Louisiana'),
    ('ME', 'Maine'),
    ('MD', 'Maryland'),
    ('MA', 'Massachusetts'),
    ('MI', 'Michigan'),
    ('MN', 'Minnesota'),
    ('MS', 'Mississippi'),
    ('MO', 'Missouri'),
    ('MT', 'Montana'),
    ('NE', 'Nebraska'),
    ('NV', 'Nevada'),
    ('NH', 'New Hampshire'),
    ('NJ', 'New Jersey'),
    ('NM', 'New Mexico'),
    ('NY', 'New York'),
    ('NC', 'North Carolina'),
    ('ND', 'North Dakota'),
    ('OH', 'Ohio'),
    ('OK', 'Oklahoma'),
    ('OR', 'Oregon'),
    ('PA', 'Pennsylvania'),
    ('RI', 'Rhode Island'),
    ('SC', 'South Carolina'),
    ('SD', 'South Dakota'),
    ('TN', 'Tennessee'),
    ('TX', 'Texas'),
    ('UT', 'Utah'),
    ('VT', 'Vermont'),
    ('VA', 'Virginia'),
    ('WA', 'Washington'),
    ('WV', 'West Virginia'),
    ('WI', 'Wisconsin'),
    ('WY', 'Wyoming'),
    ('DC', 'Washington DC')
  AS t(state_code, state_name)
""")

query_tbl_19 = f"""
SELECT
    CASE
      WHEN l.state_name IS NOT NULL THEN l.state_name
      WHEN f.ste_ctry_cd = 'DC' THEN 'Washington DC'
      ELSE 'OTHER¹'
    END AS State,
    COUNT(f.ser_num) AS Filings
FROM `{reporting_catalog}`.gold.filings_dashboard f
LEFT JOIN state_lookup l
  ON f.ste_ctry_cd = l.state_code
WHERE f.filing_fy = {current_fy}
  AND f.ctry_nm ILIKE '%UNITED STATES%'
GROUP BY
  CASE
    WHEN l.state_name IS NOT NULL THEN l.state_name
    WHEN f.ste_ctry_cd = 'DC' THEN 'Washington DC'
    ELSE 'OTHER¹'
  END
"""

df_tbl_19 = spark.sql(query_tbl_19).toPandas()

# Use the current_fy widget parameter
current_fy_19 = current_fy

# Rename column to just the year number
df_tbl_19 = df_tbl_19.rename(columns={"Filings": str(current_fy_19)})

# COMMAND ----------

# DBTITLE 1,df20
# Note: 'OTHER¹' includes Puerto Rico, Army Post Office filings, American Samoa,
# Guam, U.S. Pacific Islands, and miscellaneous U.S. State / No State Indicated in Database
with_state_lookup_20 = f"""
SELECT
    CASE
      WHEN l.state_name IS NOT NULL THEN l.state_name
      WHEN p.state = 'DC' THEN 'Washington DC'
      ELSE 'OTHER¹'
    END AS State,
    COUNT(DISTINCT p.serial_number) AS Registrations
FROM `{reporting_catalog}`.gold.post_reg_dashboard p
LEFT JOIN state_lookup l
  ON p.state = l.state_code
WHERE p.reg_fy = {current_fy}
  AND p.country_or_area_name = 'United States of America'
GROUP BY
  CASE
    WHEN l.state_name IS NOT NULL THEN l.state_name
    WHEN p.state = 'DC' THEN 'Washington DC'
    ELSE 'OTHER¹'
  END
"""

df_tbl_20 = spark.sql(with_state_lookup_20).toPandas()

# Use the current_fy widget parameter
current_fy_20 = current_fy

# Rename column to just the year number
df_tbl_20 = df_tbl_20.rename(columns={"Registrations": str(current_fy_20)})

# COMMAND ----------

# DBTITLE 1,df21
# Query from static table annual_workload_table_21 and pivot to wide format
# Get last 5 fiscal years from the static table, ensuring current year is included
last_5_filing_fys = [
    row.fiscal_year for row in spark.sql(f"""
        SELECT DISTINCT fiscal_year 
        FROM `{reporting_catalog}`.gold.annual_workload_table_21
        WHERE fiscal_year IS NOT NULL
        ORDER BY fiscal_year DESC
        LIMIT 5
    """).collect()
]

# Ensure current_fy is included in the list
if current_fy not in last_5_filing_fys:
    last_5_filing_fys.append(current_fy)
    last_5_filing_fys = sorted(last_5_filing_fys)
    # Keep only the last 5 years
    if len(last_5_filing_fys) > 5:
        last_5_filing_fys = last_5_filing_fys[-5:]
else:
    last_5_filing_fys = sorted(last_5_filing_fys)

# Build pivot query
pivot_cols_21 = ",\n".join([
    f"SUM(CASE WHEN fiscal_year = {fy} THEN application_count ELSE 0 END) AS FY{fy}"
    for fy in last_5_filing_fys
])

tbl_21_query = f"""
SELECT 
    residence AS Residence,
    {pivot_cols_21}
FROM `{reporting_catalog}`.gold.annual_workload_table_21
GROUP BY residence
HAVING {' + '.join([f'SUM(CASE WHEN fiscal_year = {fy} THEN application_count ELSE 0 END)' for fy in last_5_filing_fys])} > 0
ORDER BY 
    CASE WHEN residence = 'Other¹' THEN 1 ELSE 0 END,
    residence
"""

df2 = spark.sql(tbl_21_query).toPandas()

# Add totals and blank rows for Table 21
df2_total_row = df2.sum(numeric_only=True)
df2_total_row['Residence'] = 'Total'
total_df = pd.DataFrame([df2_total_row])
blank_row = pd.DataFrame([{col: "" for col in df2.columns}])

df2_with_total = pd.concat([
    pd.DataFrame(columns=df2.columns),
    blank_row,
    total_df,
    blank_row,
    df2
], ignore_index=True)

# Reorder columns dynamically for Table 21
fy_cols_21 = [f"FY{fy}" for fy in last_5_filing_fys]
ordered_cols_21 = ["Residence"] + fy_cols_21
df2_with_total = df2_with_total[ordered_cols_21]

# COMMAND ----------

# DBTITLE 1,df22
# Query from static table annual_workload_table_22 and pivot to wide format
# Get last 5 fiscal years from the static table
last_5_fys = [
    row.fiscal_year for row in spark.sql(f"""
        SELECT DISTINCT fiscal_year 
        FROM `{reporting_catalog}`.gold.annual_workload_table_22
        WHERE fiscal_year IS NOT NULL
        ORDER BY fiscal_year DESC
        LIMIT 5
    """).collect()
]
last_5_fys = sorted(last_5_fys)

# Build pivot query
pivot_cols = ",\n".join([
    f"SUM(CASE WHEN fiscal_year = {fy} THEN registration_count ELSE 0 END) AS FY{fy}"
    for fy in last_5_fys
])

tbl_22_query = f"""
SELECT 
    residence AS Residence,
    {pivot_cols}
FROM `{reporting_catalog}`.gold.annual_workload_table_22
GROUP BY residence
HAVING {' + '.join([f'SUM(CASE WHEN fiscal_year = {fy} THEN registration_count ELSE 0 END)' for fy in last_5_fys])} > 0
ORDER BY 
    CASE WHEN residence = 'Other¹' THEN 1 ELSE 0 END,
    residence
"""

df1 = spark.sql(tbl_22_query).toPandas()

# Add totals and blank rows for Table 22
df1_total_row = df1.sum(numeric_only=True)
df1_total_row['Residence'] = 'Total'
total_df = pd.DataFrame([df1_total_row])
blank_row = pd.DataFrame([{col: "" for col in df1.columns}])

df1_with_total = pd.concat([
    pd.DataFrame(columns=df1.columns),
    blank_row,
    total_df,
    blank_row,
    df1
], ignore_index=True)

# Reorder columns dynamically
fy_cols = [f"FY{fy}" for fy in last_5_fys]
ordered_cols = ["Residence"] + fy_cols
df1_with_total = df1_with_total[ordered_cols]

# COMMAND ----------

# MAGIC %md
# MAGIC ## Emailing

# COMMAND ----------

# DBTITLE 1,Bytestream
def create_combined_excel_bytestream(df_table_16, df_table_17, df_table_18, df_table_19, df_table_20, df_table_21, df_table_22, last_5_fys, last_5_filing_fys, current_fy_19, current_fy_20):
    """
    Create a single Excel workbook with all seven tables as separate sheets.
    
    Args:
        df_table_16: DataFrame for Table 16
        df_table_17: DataFrame for Table 17
        df_table_18: DataFrame for Table 18
        df_table_19: DataFrame for Table 19
        df_table_20: DataFrame for Table 20
        df_table_21: DataFrame for Table 21
        df_table_22: DataFrame for Table 22
        last_5_fys: List of fiscal years for Table 22 (registrations)
        last_5_filing_fys: List of fiscal years for Table 21 (filings)
        current_fy_19: Current fiscal year for Table 19
        current_fy_20: Current fiscal year for Table 20
    
    Returns:
        bytes: Excel workbook as bytes
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create Table 16 sheet
    ws_table_16 = wb.create_sheet(title="Table 16 (TM)")
    format_table_16_sheet(ws_table_16, df_table_16)

    # Create Table 17 sheet
    ws_table_17 = wb.create_sheet(title="Table 17 (TM)")
    format_table_17_sheet(ws_table_17, df_table_17)

    # Create Table 18 sheet
    ws_table_18 = wb.create_sheet(title="Table 18 (TM)")
    format_table_18_sheet(ws_table_18, df_table_18)

    # Create Table 19 sheet
    ws_table_19 = wb.create_sheet(title="Table 19 (TM)")
    format_state_table_sheet(ws_table_19, df_table_19, "19", 
                            "TRADEMARK APPLICATIONS FILED BY RESIDENTS OF THE UNITED STATES", 
                            current_fy_19)

    # Create Table 20 sheet
    ws_table_20 = wb.create_sheet(title="Table 20 (TM)")
    format_state_table_sheet(ws_table_20, df_table_20, "20", 
                            "TRADEMARKS REGISTERED TO RESIDENTS OF THE UNITED STATES", 
                            current_fy_20)

    # Create Table 21 sheet
    ws_table_21 = wb.create_sheet(title="Table 21 (TM)")
    format_foreign_country_sheet(ws_table_21, df_table_21, "Table 21", last_5_filing_fys)

    # Create Table 22 sheet
    ws_table_22 = wb.create_sheet(title="Table 22 (TM)")
    format_foreign_country_sheet(ws_table_22, df_table_22, "Table 22", last_5_fys)

    # Save to BytesIO and return bytes
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.getvalue()

# COMMAND ----------

# DBTITLE 1,Email Function
def send_combined_email(
    send_from: str,
    send_to: str,
    send_to_cc: str,
    subject: str,
    text: str,
    df_table_16,
    df_table_17,
    df_table_18,
    df_table_19,
    df_table_20,
    df_table_21,
    df_table_22,
    last_5_fys,
    last_5_filing_fys,
    current_fy_19,
    current_fy_20,
    attachment_name: str = None,  # Defaults to "TM_{current_fy}_Annual_Workload_Tables.xlsx"
    server: str = "mailer.uspto.gov",
):
    """
    Send an email with all seven tables in a single Excel workbook.

    Args:
        send_from: Sender email address
        send_to: Comma-separated recipient email addresses
        send_to_cc: Comma-separated CC email addresses
        subject: Email subject
        text: Email body text
        df_table_16: DataFrame for Table 16
        df_table_17: DataFrame for Table 17
        df_table_18: DataFrame for Table 18
        df_table_19: DataFrame for Table 19
        df_table_20: DataFrame for Table 20
        df_table_21: DataFrame for Table 21
        df_table_22: DataFrame for Table 22
        last_5_fys: List of fiscal years for Table 22 (registrations)
        current_fy_19: Current fiscal year for Table 19
        current_fy_20: Current fiscal year for Table 20
        attachment_name: Name of the Excel file attachment (e.g. f"TM_{current_fy}_Annual_Workload_Tables.xlsx"). Defaults to that pattern if not provided.
        server: SMTP server address
    """
    if attachment_name is None:
        attachment_name = f"TM_Annual_Workload_Tables.xlsx"  # Pass current_fy explicitly for a versioned filename

    try:
        msg = MIMEMultipart()
        msg["From"] = send_from
        msg["To"] = COMMASPACE.join(send_to.split(","))
        msg["Cc"] = COMMASPACE.join(send_to_cc.split(",")) if send_to_cc else ""
        msg["Subject"] = subject

        msg.attach(MIMEText(text))

        # Create the combined Excel file
        excel_bytes = create_combined_excel_bytestream(
            df_table_16,
            df_table_17,
            df_table_18,
            df_table_19,
            df_table_20,
            df_table_21,
            df_table_22,
            last_5_fys,
            last_5_filing_fys,
            current_fy_19,
            current_fy_20,
        )

        part = MIMEApplication(excel_bytes)
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            "attachment",
            filename=attachment_name,
        )
        msg.attach(part)

        smtp = smtplib.SMTP(server)
        rcpt = send_to.split(",") + (send_to_cc.split(",") if send_to_cc else [])
        smtp.sendmail(send_from, rcpt, msg.as_string())
        smtp.close()

        print(f"Email sent successfully to {send_to}")

    except Exception as error:
        print("An issue occurred during the email sending process.")
        raise error

# COMMAND ----------

# DBTITLE 1,Send Combined Email
print("Sending email...")
send_combined_email(
    send_from="trademark_analytics@uspto.gov",
    send_to= primary_email,
    send_to_cc= cc_email,
    subject="Annual Workload Tables",
    text=f"Please find attached the Trademark Annual Workload Report for FY{current_fy} to date.",
    df_table_16=df_tbl_16,
    df_table_17=df_tbl_17,
    df_table_18=df_tbl_18,
    df_table_19=df_tbl_19,
    df_table_20=df_tbl_20,
    df_table_21=df2_with_total,
    df_table_22=df1_with_total,
    last_5_fys=last_5_fys,
    last_5_filing_fys=last_5_filing_fys,
    current_fy_19=current_fy_19,
    current_fy_20=current_fy_20,
    attachment_name=f"TM_{current_fy}_Annual_Workload_Tables.xlsx",
)

# COMMAND ----------

# DBTITLE 1,End Job Control
# Count total tables generated
recs_count = 7  # 7 tables: 16, 17, 18, 19, 20, 21, 22

end_job_cntl(
    f"{reporting_catalog}.silver",
    job_name,
    job_start_ts,
    "completed",
    recs_count,
    "AFR Combined Tables Email Report completed successfully",
)

dbutils.notebook.exit(f"Completed AFR Combined Tables Email Report - {recs_count} tables generated and emailed")