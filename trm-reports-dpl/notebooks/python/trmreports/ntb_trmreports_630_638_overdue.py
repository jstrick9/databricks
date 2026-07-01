# Databricks notebook source
from pyspark.sql.functions import *

# COMMAND ----------

dbutils.widgets.text("dbx_env","dev")

# COMMAND ----------

dbx_env = dbutils.widgets.get("dbx_env").rstrip()
config_file_name = "trmreports-conf.yaml"
config_file = "../../config/"+dbutils.widgets.get("dbx_env")+"/"+config_file_name
print(f'{config_file=}')

# COMMAND ----------

# MAGIC %run ../shared/ntb_common_func_and_params $config_file=config_file 

# COMMAND ----------

common_configs = read_yaml(config_file)
reporting_catalog = common_configs['schema']['trgt_catalog']
tmngpdb_catalog = common_configs['schema']['tmngpdb_src_catalog']
tmworker_catalog = common_configs['schema']['tmworker_catalog']
to_email_address = common_configs['alerting']['630_638_overdue']['email']
altrx_schema = common_configs['schema']['altrx_schema']
dq_catalog = common_configs['schema']['data_quality_catalog']

# COMMAND ----------

# DBTITLE 1,Start Job Control
# set current time for both while loop and job control
curntdt = datetime.datetime.now().astimezone(pytz.timezone('US/Eastern'))

# start job control  
starttime = curntdt.strftime('%Y-%m-%d %H:%M:%S')
job_name = 'ntb_trmreports_630_638_overdue'

control_dt = begin_job_cntl(f'{reporting_catalog}.silver',job_name,starttime)

# COMMAND ----------

# MAGIC %md
# MAGIC %md
# MAGIC #### ETL

# COMMAND ----------

DF_381 = spark.sql(f"Select * FROM {reporting_catalog}.gold.pendency_dashboard")


# COMMAND ----------

DF_31 = DF_381.filter(col("first_action_dt_ph").isNotNull())


# COMMAND ----------

DF_21 = DF_31.select(
    "first_action_pendency_ph",
    "first_action_dt_ph",
    "first_action_type_num",
    "fa_pendency_filter",
    "filing_basis_grp",
    "pendency_cal_end_dt",
    "ser_num",
)


# COMMAND ----------

DF_7 = DF_21.withColumn(
    "Filing_Basis_Group",
    when(col("filing_basis_grp") == "MADRID", "MADRID").otherwise("NON-MADRID"),
)


# COMMAND ----------

DF_9T = DF_7.filter(col("filing_basis_group") == "MADRID")
DF_9F = DF_7.filter(col("filing_basis_group") != "MADRID")


# COMMAND ----------

DF_48 = DF_9T.withColumn(
    "Diff_Today_FirstA", (date_diff(current_date(), "first_action_dt_ph"))
)


# COMMAND ----------

DF_49 = DF_48.filter(col("Diff_Today_FirstA") <=30)


# COMMAND ----------

DF_35 = DF_49.groupby().avg("first_action_pendency_ph").withColumnRenamed(
    "avg(first_action_pendency_ph)", "AvgP_CurrentMonth_Madrid"
)

AvgP_CurrentMonth_Madrid=DF_35.collect()[0][0]


# COMMAND ----------

DF_382 = spark.sql(
    f"Select * FROM {reporting_catalog}.silver.milestone WHERE YEAR(filing_dt) > 2010 and IsNull(first_action_dt_ph)"
)


# COMMAND ----------

DF_4 = DF_382.select(
    "ser_num", 
    "first_action_dt_ph", 
    "pendency_cal_start_dt", 
    "dock_dt"
).withColumnRenamed("dock_dt", "Assigned_DT")


# COMMAND ----------

DF_29 = DF_4.filter(col("first_action_dt_ph").isNull())


# COMMAND ----------

DF_383 = spark.sql(f"Select * FROM {reporting_catalog}.silver.bibliography")


# COMMAND ----------

DF_25 = DF_383.withColumn(
   "Filing_Basis_Group", when(col("filing_basis_grp")=="MADRID","MADRID").otherwise("NON-MADRID") 
)


# COMMAND ----------

DF_24 = DF_25.select(
    "SER_NUM",
    "AM_STAT",
    "FILING_BASIS_GRP",
    "MARK_NM",
    "Filing_Basis_Group"
).withColumnRenamed("MARK_NM", "Mark_NM")


# COMMAND ----------

DF_27 = DF_24.filter((col("AM_STAT")=="630") | (col("AM_STAT")=="638"))


# COMMAND ----------

DF_28 = DF_29.join(
    DF_27,
    "SER_NUM",
    "Inner").drop("FILING_BASIS_GRP")


# COMMAND ----------

DF_390 = spark.sql(
    f"Select substr(trademark_gid,INSTR(trademark_gid,':')+3) AS AM_SER_NUM, legacy_status_cd as AM_STATUS_CD, status_dt as AM_STATUS_DT FROM {tmngpdb_catalog}.bronze.trademark"
)


# COMMAND ----------

DF_271 = DF_390.select(
  "AM_SER_NUM",
  "AM_STATUS_CD",
  "AM_STATUS_DT"
)


# COMMAND ----------

DF_270 = DF_28.join(
    DF_271,
    DF_28["ser_num"] == DF_271["AM_SER_NUM"],
    "inner"
).withColumnRenamed("STATUS_DT","Right_STATUS_DT")
DF_270 = DF_270.drop(
    "AM_SER_NUM")


# COMMAND ----------

DF_39 = DF_270.withColumn(
    "LivePendency_NoFirstAction",
    datediff(current_date(), col("Pendency_Cal_Start_DT")) / 30.42
).withColumn(
    "Assigned Days",
    expr("datediff(current_date(), Assigned_DT)"))



# COMMAND ----------

DF_32T = DF_39.filter(col("Filing_Basis_Group") == "MADRID")
DF_32F = DF_39.filter(col("Filing_Basis_Group") != "MADRID")


# COMMAND ----------

DF_41 = DF_32T.withColumn(
   "AvgP_CurrentMonth_Madrid", lit(AvgP_CurrentMonth_Madrid)
)


# COMMAND ----------

DF_42 = DF_41.withColumn("Delta", col("LivePendency_NoFirstAction") - col("AvgP_CurrentMonth_Madrid"))

DF_42 = DF_42.drop("AM_SER_NUM")


# COMMAND ----------

DF_44 = DF_42.filter((col("Delta")>=2) | (col("Assigned Days")>=20))


# COMMAND ----------

DF_46 = DF_44.orderBy(col("Delta").desc())
DF_42 = DF_42.drop("AM_SER_NUM")


# COMMAND ----------

DF_384 = spark.sql(f"""
    SELECT *,
        CAST(split(wio.cfk_object_gid, ':')[2] AS STRING) AS SER_NUM
    FROM
        {tmngpdb_catalog}.bronze.attorney_hold ath
    JOIN
         {tmngpdb_catalog}.bronze.work_item_object wio
    ON
        ath.fk_work_item_gid = wio.fk_work_item_gid
""")

DF_384 = DF_384.drop("fk_work_item_gid2",
                      "fk_object_type_cd",
                      "cfk_object_gid",
                      "lock_control_no2",
                      "create_ts2",
                      "create_user_id2",
                      "last_mod_user_id2")


# COMMAND ----------

DF_385 = DF_384.select(
    DF_384["ath.PLACED_ON_HOLD_DT"],
    DF_384["ath.FK_WORK_ITEM_GID"],
    DF_384["ath.CFK_HOLD_WORKER_GID"],
    DF_384["ath.CFK_HOLD_STATUS_CD"],
    DF_384["ath.HOLD_DOCKET_NO"],
    DF_384["ath.DN_SERIAL_NUM_TX"],
    DF_384["ath.LAST_ACTION_DT"],
    DF_384["ath.CFK_HOLD_USER_ROLE_ID"],
    DF_384["ath.CFK_HOLD_TM_ORGANIZATION_GID"],
    DF_384["ath.CFK_HOLD_CATEGORY_CD"],
    DF_384["ath.CFK_LAST_ACTION_WORKER_GID"],
    DF_384["ath.CFK_LAST_ACTION_USER_ROLE_ID"],
    DF_384["ath.CFK_LAST_ACTION_TM_ORG_GID"],
    DF_384["ath.LOCK_CONTROL_NO"],
    DF_384["ath.CREATE_TS"],
    DF_384["ath.LAST_MOD_TS"],
    DF_384["ath.CREATE_USER_ID"],
    DF_384["ath.LAST_MOD_USER_ID"],
    DF_384["ath.DN_HOLD_WORKER_NO"],
    DF_384["ath.DN_LAST_ACTION_WORKER_NO"],
    DF_384["SER_NUM"]
).withColumnsRenamed(
    {"DN_SERIAL_NUM_TX": "ATH_SER_NUM",
     "CFK_HOLD_WORKER_GID": "ATH_EMPE_NUM",
     "CFK_HOLD_STATUS_CD": "ath_hold_status",
     "HOLD_DOCKET_NO": "ATH_HOLD_DOCKET",
     "PLACED_ON_HOLD_DT": "ATH_CREATE_DT"}
)


# COMMAND ----------

DF_53 = DF_385.select(col("ath_ser_num").alias("ser_num"), "ath_hold_status")

# COMMAND ----------

DF_268 = DF_46.join(
    DF_53,
    'ser_num',
    "left"
)

# COMMAND ----------

from pyspark.sql.functions import round

DF_57 = DF_268.select(
    DF_268["ser_num"],
    DF_268["first_action_dt_ph"],
    DF_268["pendency_cal_start_dt"],
    DF_268["Assigned_DT"],
    DF_268["AM_STAT"],
    DF_268["Mark_NM"],
    DF_268["Filing_Basis_Group"],
    DF_268["AM_STATUS_CD"],
    DF_268["AM_STATUS_DT"],
    DF_268["Assigned Days"],
    DF_268["LivePendency_NoFirstAction"],
    "ath_hold_status"
).withColumnRenamed("AM_STAT", "Status") \
 .withColumnRenamed("Filing_Basis_Group", "Basis") \
 .withColumn("Age Months", round(DF_268["LivePendency_NoFirstAction"], 1))


# COMMAND ----------

DF_50 = DF_9F.withColumn(
    "Diff_Today_FirstA", (date_diff(current_date(), "first_action_dt_ph"))
)


# COMMAND ----------

DF_51 = DF_50.filter(col("Diff_Today_FirstA") <=30)


# COMMAND ----------

DF_38 = DF_51.groupby().avg("first_action_pendency_ph").withColumnRenamed(
    "avg(first_action_pendency_ph)", "AvgP_CurrentMonth_NonMadrid"
)

AvgP_CurrentMonth_Madrid=DF_38.collect()[0][0]



# COMMAND ----------

DF_40 = DF_32F.withColumn(
   "AvgP_CurrentMonth_NonMadrid", lit(AvgP_CurrentMonth_Madrid)
)


# COMMAND ----------

DF_43 = DF_40.withColumn("Delta", col("LivePendency_NoFirstAction") - col("AvgP_CurrentMonth_NonMadrid"))

DF_43 = DF_43.drop("AM_SER_NUM")


# COMMAND ----------

DF_45 = DF_43.filter((col("Delta")>=2) | (col("Assigned Days")>=20))


# COMMAND ----------

DF_47 = DF_45.orderBy(col("Delta").desc())


# COMMAND ----------

DF_55 = DF_47.join(
    DF_53,
    "ser_num",
    "left"
)

DF_55.count()

# DF_268 = DF_46.join(
#     DF_53,
#     DF_46['ser_num'] == DF_53['ATH_SER_NUM'],
#     "left_anti",
# ) 


# COMMAND ----------

DF_58 = DF_55.select(
    DF_268["ser_num"],
    DF_268["first_action_dt_ph"],
    DF_268["pendency_cal_start_dt"],
    DF_268["Assigned_DT"],
    DF_268["AM_STAT"],
    DF_268["Mark_NM"],
    DF_268["Filing_Basis_Group"],
    DF_268["AM_STATUS_CD"],
    DF_268["AM_STATUS_DT"],
    DF_268["Assigned Days"],
    DF_268["LivePendency_NoFirstAction"],
    "ath_hold_status"
).withColumnRenamed("AM_STAT", "Status") \
 .withColumnRenamed("Filing_Basis_Group", "Basis") \
 .withColumn("Age Months", round(DF_268["LivePendency_NoFirstAction"], 1))



# COMMAND ----------

DF_60 = DF_57.union(DF_58) \
 .withColumnRenamed("ser_num","Serial Number") \
 .withColumnRenamed("Pendency_Cal_Start_DT","Filing Date")


# COMMAND ----------

DF_394 = DF_383.select(
    "SER_NUM",
    "TEST_PCTRAM_LINK",
    "LAW_OFFICE",
    "FILING_BASIS_CUR",
    "FILING_METHOD_FILED",
    "FILING_METHOD_CUR",
    "FILING_BASIS_FIL",
    "FILING_BASIS_AMED",
    "REGISTRATION_NUMBER",
    "AM_FLG_66A_FIL",
    "AM_FLG_44D_FIL",
    "AM_FLG_44E_FIL",
    "FLG_PAPER_FIL",
    "AM_STAT",
    "AM_FLG_NO_BAS_FIL",
    "AM_FLG_TEASRF_FIL",
    "AM_FLG_USE_FIL",
    "AM_FLG_ITU_FIL",
    "AM_FLG_TEASPL_FIL",
    "LAST_MODIFIED_DATE",
    "FILING_BASIS_GRP",
    "MARK_DWG_CD",
    "MARK_DWG_DESC",
    "MARK_NM_SHORT",
    "MARK_NM",
    "TMNG_IMAGE_LINK",
    "TM_ANALYTICS_TS",
    "EXMR_EID",
    "STATUS_DT",
    "create_ts",
    "create_user_id",
    "update_ts",
    "update_user_id"
).withColumnRenamed("STATUS_DT", "STATUS_DT2")
DF_394 = DF_394.drop(
    "FILING_BASIS_CUR",
    "FILING_METHOD_FILED",
    "FILING_METHOD_CUR",
    "FILING_BASIS_FIL",
    "FILING_BASIS_AMED",
    "REGISTRATION_NUMBER",
    "AM_FLG_66A_FIL",
    "AM_FLG_44D_FIL",
    "AM_FLG_44E_FIL",
    "FLG_PAPER_FIL",
    "AM_STAT",
    "AM_FLG_NO_BAS_FIL",
    "AM_FLG_TEASRF_FIL",
    "AM_FLG_USE_FIL",
    "AM_FLG_ITU_FIL",
    "AM_FLG_TEASPL_FIL",
    "LAST_MODIFIED_DATE",
    "FILING_BASIS_GRP",
    "MARK_DWG_CD",
    "MARK_DWG_DESC",
    "MARK_NM",
    "TMNG_IMAGE_LINK",
    "TM_ANALYTICS_TS")


# COMMAND ----------

DF_61 = DF_60.join(
    DF_394,
    DF_60["Serial Number"] == DF_394["SER_NUM"],
    "inner"
)
DF_61 = DF_61.drop("SER_NUM")


# COMMAND ----------

DF_387 = spark.sql(f"""
SELECT w.worker_no as EE_EMPE_NUM,
w.worker_nm as EE_EMPE_NAM,
org.organization_cd as EE_EMPE_LO,
org.organization_nm,
wr.begin_effective_dt,
wr.end_effective_dt
FROM {tmworker_catalog}.bronze.worker w
JOIN {tmworker_catalog}.bronze.worker_role wr ON w.worker_gid = wr.fk_worker_gid
JOIN {tmworker_catalog}.bronze.tm_organization org on wr.fk_tm_organization_gid = org.tm_organization_gid
ORDER BY worker_no
""")


# COMMAND ----------

DF_63 = DF_61.join(DF_387, DF_61["EXMR_EID"] == DF_387["EE_EMPE_NUM"], "left_outer")

DF_63 = DF_63.drop(
"EE_EMPE_LO")


# COMMAND ----------

DF_391 = spark.sql(
    f"""
    SELECT 
        ph_action_code,
        CONCAT(SUBSTRING(ph_action_code, 1, 4), fifth_char_cm_type) AS 5CharCode,
        CASE 
            WHEN MONTH(cm_sys_dt) > 9 THEN YEAR(cm_sys_dt) + 1 
            ELSE YEAR(cm_sys_dt) 
        END AS FY,
        serial_number,
        create_ts,
        ph_action_number,
        ph_action_date
    FROM {reporting_catalog}.silver.prosecution_history
    WHERE CONCAT(SUBSTRING(ph_action_code, 1, 4), fifth_char_cm_type) = 'DOCKD'
    ORDER BY serial_number, ph_action_number
    """
)


# COMMAND ----------

DF_392 = (DF_391.select(
    "ph_action_code",
    "5CharCode",
    "FY",
    "serial_number",
    "create_ts",
    "ph_action_number",
    "ph_action_date"
).withColumnRenamed("serial_number", "CM_SER_NUM")
 .withColumnRenamed("ph_action_number", "CM_ENT_NUM"))



# COMMAND ----------

DF_274 = DF_392.groupby("CM_SER_NUM").agg(max("CM_ENT_NUM").alias("Max_CM_ENT_NUM"))

DF_274 = DF_274.drop(
    "ph_action_code"
)


# COMMAND ----------

DF_275 = DF_274.join(
    DF_392,
    (DF_274.CM_SER_NUM == DF_392.CM_SER_NUM)
    & (DF_274.Max_CM_ENT_NUM == DF_392.CM_ENT_NUM), "inner"
).drop(DF_392.CM_SER_NUM)

DF_275 = DF_275.drop(
    "Max_CM_ENT_NUM"
    "ph_action_code",
    "5CharCode",
    "FY",
    "create_ts",
    "CM_ENT_NUM",
    "ph_action_date",
).withColumnRenamed("CM_ENT_DT", "Current Assigned Date")



# COMMAND ----------

DF_276 = DF_63.join(DF_275, DF_63["Serial Number"] == DF_275["CM_SER_NUM"],"left_outer")


# COMMAND ----------

DF_91 = DF_276.orderBy(col("Age Months").desc())


# COMMAND ----------

DF_65 = DF_91.select(
    DF_91["Serial Number"],
    DF_91["EE_EMPE_NAM"],
    DF_91["Mark_NM"],
    DF_91["Filing Date"],
    DF_91["Age Months"],
    DF_91["Assigned_DT"],
    DF_91["Assigned Days"],
    DF_91["Status"],
    DF_91["Basis"],
    DF_91["LAW_OFFICE"],
    DF_91["first_action_dt_ph"],
    DF_91["AM_STATUS_CD"],
    DF_91["AM_STATUS_DT"],
    DF_91["TEST_PCTRAM_LINK"],
    DF_91["MARK_NM_SHORT"],
    DF_91["STATUS_DT2"],
    DF_91["create_ts"],
    DF_91["create_user_id"],
    DF_91["update_ts"],
    DF_91["update_user_id"],
    DF_91["organization_nm"],
    DF_91["AM_STATUS_CD"],
    DF_91["BEGIN_EFFECTIVE_DT"],
    DF_91["END_EFFECTIVE_DT"],
    DF_91["EXMR_EID"],
    "ath_hold_status"
).withColumnRenamed("EE_EMPE_NAM", "Examiner") \
 .withColumnRenamed("Mark_NM", "Mark") \
 .withColumnRenamed("Filing Date", "filing_ib_date") \
 .withColumnRenamed("Assigned_DT", "Original Assigned Date") \
 .withColumnRenamed("Assigned Days", "Original Assigned Days") \
 .withColumnRenamed("LAW_OFFICE", "Law Office")


# COMMAND ----------

from pyspark.sql.functions import col, last
from pyspark.sql.window import Window

# COMMAND ----------

win396 = Window().partitionBy(
    "Serial Number",
    "Examiner",
    "Mark",
    "filing_ib_date",
    "Age Months",
    "Original Assigned Date",
    "Original Assigned Days",
    "Status",
    "Basis",
    "Law Office",
    "first_action_dt_ph",
    "AM_STATUS_DT",
    "TEST_PCTRAM_LINK",
    "MARK_NM_SHORT"
).orderBy(
    col('EXMR_EID'),
    col("Age Months"),
    col("BEGIN_EFFECTIVE_DT")
)
 
DF_396 = DF_65.withColumn(
    "last_organization_nm", first("organization_nm").over(win396)
).select("Serial Number",
    "Examiner",
    "Mark",
    "filing_ib_date",
    "Age Months",
    "Original Assigned Date",
    "Original Assigned Days",
    "Status",
    "Basis",
    "Law Office",
    "first_action_dt_ph",
    "AM_STATUS_DT",
    "TEST_PCTRAM_LINK",
    "MARK_NM_SHORT",
    "last_organization_nm",
    "ath_hold_status"
    ).distinct()


# COMMAND ----------

DF_66 = DF_396.withColumn(
    "Law Office", 
    when(col("Law Office").isNull(), lit('N/A')).otherwise(col("Law Office"))
).withColumn(
    "ContentLink", 
    concat(lit('https://review.tm-examcenter.aws.uspto.gov/review/'), col("Serial Number"))
).withColumn(
    "PCTRAMName", lit("PCTRAM")
).withColumn(
    "ContentName", lit("Content Mgr")
)   

# COMMAND ----------

# MAGIC %md
# MAGIC ### **OUTPUT**
# MAGIC

# COMMAND ----------

# MAGIC %pip install openpyxl

# COMMAND ----------

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill
import shutil

# COMMAND ----------

# COMMAND ----------
# DBTITLE 1,Prepare Output DataFrame with Correct Types and Column Names

from pyspark.sql.functions import col, lit, when, to_date, round as spark_round
from pyspark.sql.types import IntegerType, DecimalType, BooleanType, DateType

df_out = DF_66.select(
    # String columns - OK as is
    col('Serial Number').cast(StringType()).alias('serial_num'),
    col('Examiner').cast('string').alias('Examiner'),
    col('ContentName').cast('string').alias('ContentManager'),
    col('Mark').cast('string').alias('Mark'),
    
    # Date columns - cast to date type
    to_date(col('filing_ib_date')).alias('Filing_IB_Date'),
    
    # Decimal column - cast appropriately
    spark_round(col('Age Months'), 0).cast(DecimalType(10, 0)).alias('Age_Months'),
    
    # Date column
    to_date(col('Original Assigned Date')).alias('Original_Assigned_Date'),
    
    # Integer column
    col('Original Assigned Days').cast(IntegerType()).alias('Original_Assigned_Days'),
    
    # Status - cast to integer (table expects int, not string!)
    col('Status').cast(IntegerType()).alias('Status'),
    
    # String columns
    col('Basis').cast('string').alias('Basis'),
    col('Law Office').cast('string').alias('Law_Office'),
    col('first_action_dt_ph').cast('string').alias('first_action_dt_ph'),
    
    # Date column
    to_date(col('AM_STATUS_DT')).alias('AM_STATUS_DT'),
    
    # String columns
    col('MARK_NM_SHORT').cast('string').alias('MARK_NM_SHORT'),
    col('last_organization_nm').cast('string').alias('Last_organization_nm'),
    col('ContentLink').cast('string').alias('ContentLink'),
    
    # CRITICAL FIX: Rename ath_hold_status to on_hold and convert to boolean
    # If ath_hold_status contains values like 'ASSIGNED_EXAMINER', 'Y', 'N', etc.
    # we need to convert to boolean
    when(
        col('ath_hold_status').isNull(), lit(None).cast(BooleanType())
    ).when(
        col('ath_hold_status').isin(['Y', 'YES', 'TRUE', '1', 'HOLD']), lit(True)
    ).when(
        col('ath_hold_status').isin(['N', 'NO', 'FALSE', '0', 'ASSIGNED_EXAMINER']), lit(False)
    ).otherwise(
        lit(False)  # Default to False for any other value
    ).alias('on_hold')
)
 
all_los = sorted([x['Law_Office'] for x in df_out.select("Law_Office").distinct().collect()])

# COMMAND ----------

df_out_hist = df_out.withColumn(
    'create_ts', from_utc_timestamp(current_timestamp(), 'US/Eastern')
)

# COMMAND ----------

# overwrite main table
df_out.write.mode("overwrite").format("delta").insertInto(f"{reporting_catalog}.gold.overdue_630_638")

# merge to hist table using create_ts date for same day re-runs
df_out_hist.createOrReplaceTempView("df_out_hist")

spark.sql(f"""
MERGE INTO {reporting_catalog}.gold.overdue_630_638_hist existing
USING df_out_hist new
ON existing.serial_num = new.serial_num AND to_date(existing.create_ts) = to_date(new.create_ts)
WHEN MATCHED THEN UPDATE SET
    serial_num = new.serial_num,
    Examiner = new.Examiner,
    ContentManager = new.ContentManager,
    Mark = new.Mark,
    Filing_IB_Date = new.Filing_IB_Date,
    Age_Months = new.Age_Months,
    Original_Assigned_Date = new.Original_Assigned_Date,
    Original_Assigned_Days = new.Original_Assigned_Days,
    Status = new.Status,
    Basis = new.Basis,
    Law_Office = new.Law_Office,
    first_action_dt_ph = new.first_action_dt_ph,
    AM_STATUS_DT = new.AM_STATUS_DT,
    MARK_NM_SHORT = new.MARK_NM_SHORT,
    Last_organization_nm = new.Last_organization_nm,
    ContentLink = new.ContentLink,
    on_hold = new.on_hold,
    create_ts = new.create_ts
WHEN NOT MATCHED THEN INSERT (
    serial_num,
    Examiner,
    ContentManager,
    Mark,
    Filing_IB_Date,
    Age_Months,
    Original_Assigned_Date,
    Original_Assigned_Days,
    Status,
    Basis,
    Law_Office,
    first_action_dt_ph,
    AM_STATUS_DT,
    MARK_NM_SHORT,
    Last_organization_nm,
    ContentLink,
    on_hold,
    create_ts
) VALUES (
    new.serial_num,
    new.Examiner,
    new.ContentManager,
    new.Mark,
    new.Filing_IB_Date,
    new.Age_Months,
    new.Original_Assigned_Date,
    new.Original_Assigned_Days,
    new.Status,
    new.Basis,
    new.Law_Office,
    new.first_action_dt_ph,
    new.AM_STATUS_DT,
    new.MARK_NM_SHORT,
    new.Last_organization_nm,
    new.ContentLink,
    new.on_hold,
    new.create_ts
)
""").display()


# COMMAND ----------

# MAGIC %md
# MAGIC #### Send Email Reports

# COMMAND ----------

## read back from tables before email output for better performance
df_main = spark.sql(f"select * from {reporting_catalog}.gold.overdue_630_638").withColumn(
    'age_months', col('age_months').cast(StringType())
).withColumnsRenamed(
    {'law_office': 'Law Office',
     'serial_num': 'Serial Number',
     'filing_ib_date' : 'Filing/IB Date',
     'age_months': 'Age Months',
     'Original_Assigned_Date' : 'Original Assigned Date',
     'Original_Assigned_Days' : 'Original Assigned Days'}
)

df_hist = spark.sql(f"select law_office as `Law Office`, serial_num, date_format(create_ts, 'y-MM') as Month from {reporting_catalog}.gold.overdue_630_638_hist")

# COMMAND ----------

def set_links(ws):
 
    ### Identify columns 
    Content_col_letter = ''
    Content_link_col_letter = ''
    Content_link_col_num = 0
 
    ## find column positions for content manager and content link columns
    for r in ws.iter_rows(max_row=1):
        for c in r:
            if c.value == 'ContentManager':
                Content_col_letter = c.column_letter
            elif c.value == 'ContentLink':
                Content_link_col_letter = c.column_letter
                Content_link_col_num = c.column
 
    ## set hyperlink styling and link
    for cell in ws[Content_col_letter][1:]:
        cell.style = "Hyperlink"
        cell.hyperlink = ws[Content_link_col_letter][cell.row - 1].value
 
    ## delete content link column after setting hyperlinks
    ws.delete_cols(Content_link_col_num)
 
    return ws
 
## Apply shading to headers using pattern header_pattern.
def fill_headers(ws, header_pattern):
    for cell in ws[1]:
        cell.fill = header_pattern
    return ws
 
## Apply row banding to provide worksheet ws using color fill from banding_pattern. Apply to either even rows or odd rows using even_odd.
def row_banding(ws, banding_pattern, even_odd):
    if even_odd == 'even':
        start = 2
    elif even_odd == 'odd':
        start = 3
    else:
        return "Invalid input for even_odd"
    for row_num in range(start, ws.max_row + 1, 2):
        row = ws[row_num]
        for cell in row:
            cell.fill = banding_pattern
 
    return ws

## define simple max to avoid using python max function conflicting with pyspark max function
def mymax(a, b):
    if a > b:
        return a
    else:
        return b

## Dynamically size column widths
def set_col_widths(ws):
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = mymax(dims.get(cell.column_letter, 0), len(str(cell.value)))
    for cl, value in dims.items():
        ws.column_dimensions[cl].width = value * 1.1

    return ws

## Combine functions for single formatting master function
def format_sheet(ws):
    ws = set_col_widths(ws)
 
    # set fills
    pattern_fill_header = PatternFill(start_color='dbdbdb', end_color='dbdbdb', fill_type='solid')
    pattern_fill_banding = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
 
    ws = fill_headers(ws, pattern_fill_header)
    ws = row_banding(ws, pattern_fill_banding, 'odd')

    return ws

# COMMAND ----------

wb = openpyxl.Workbook()

## set first sheet as all LO summary over time
df_hist_pd = df_hist.groupBy("Law Office", "Month").agg(count("serial_num").alias("Cases")).orderBy("Month", "Law Office").toPandas().reset_index(drop=True)

wb.create_sheet('Summary', 0)
ws_sum = wb['Summary']

for r in dataframe_to_rows(df_hist_pd, index=False, header=True):
    ws_sum.append(r)

ws_sum = format_sheet(ws_sum)

## loop through LOs to populate remaining sheets

index = 1

for lo in all_los:
    print(f"Formatting sheet for law office {lo}")

    df = df_main.filter(col("Law Office") == lo).toPandas().reset_index(drop=True)
 
    if lo == 'N/A':
        lo = 'N_A'
 
    wb.create_sheet(lo, index)
    ws = wb[lo]
 
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    ws = set_links(ws)
    ws = format_sheet(ws)
 
    index += 1

# COMMAND ----------

import os

try:
    wb.remove(wb['Sheet'])
except KeyError:
    print("Default sheet already removed")
 
local_file_output_path = f"/tmp/630_638_overdue rprt.xlsx"
dbfs_file_output_path = f"/dbfs/mnt/eds/trademark/dbx_reports/630_638_overdue rprt.xlsx"
 
wb.save(local_file_output_path)  # Save the workbook to local path
if not os.path.exists('/dbfs/mnt/eds/trademark/dbx_reports/'):
    os.makedirs('/dbfs/mnt/eds/trademark/dbx_reports/')
shutil.move(local_file_output_path, dbfs_file_output_path)  # Move file to DBFS path


# COMMAND ----------

# Email credentials
from_addr = "Trademark_Analytics@uspto.gov"
subj = f"630 638 Overdue Report"
html = f"""
        The attached 630 638 Overdue Report calculates the following data points.<br><br>
        1. A rolling 30 day first action pendency for both madrid and non-madrid.<br>
        2. Flag all 630 and 638 cases with no first action that are 2+ months older than the 30 day rolling first action pendency (separately for madrid and non madrid).
        """
 
# Attach the PDF file
attachments = [dbfs_file_output_path]  

# Send the email with the attachment
send_email_report(
    job_nm = job_name,
    subject = subj,
    send_from = from_addr,
    send_to = to_email_address,
    html_body= html,
    attachments = attachments
)

# COMMAND ----------

#######################################################################################################################
# 5/9/25 - Commented out data quality check code following updates to the report logic that aren't in Alteryx version #
#######################################################################################################################

# data quality entry
# tbl1 = f"{reporting_catalog}.gold.overdue_630_638"
# tbl2 = f"hive_metastore.{altrx_schema}.overdue_630_638"
# key_cols = ['serial_number']

# dq_result = alteryx_data_match(tbl1, tbl2, key_cols, job_name, dq_catalog)
# print(dq_result)

# COMMAND ----------

#end job control
recs_count = df_out.count()
end_job_cntl(f"{reporting_catalog}.silver", job_name, starttime,'completed', recs_count,"job completed successfully")
