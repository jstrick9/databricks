# Databricks notebook source
from pyspark.sql.functions import col, sum as _sum, row_number, lit, when, concat_ws, concat, round, format_number
from pyspark.sql.window import Window
import pandas as pd
import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.drawing.image import Image
import boto3
import os
from datetime import datetime

# COMMAND ----------

# Get the current date
current_date = datetime.now()

# Determine the current fiscal year
# Fiscal year starts in October, so if the current month is October or later, the fiscal year is the current year + 1
if current_date.month >= 10:
    default_fiscal_year = current_date.year + 1
else:
    default_fiscal_year = current_date.year

# Set the fiscal year widget, defaulting to the calculated fiscal year if not manually provided in the workflow
dbutils.widgets.text("fiscal_year", str(default_fiscal_year))
fiscal_year_workflow = dbutils.widgets.get("fiscal_year")
fiscal_year = fiscal_year_workflow if fiscal_year_workflow.strip() != "" else str(default_fiscal_year)

# Set the environment widget
dbutils.widgets.text("dbx_env", "dev")
dbx_env = dbutils.widgets.get("dbx_env")

# Configuration file path
config_file_name = "trmreports-conf.yaml"
config_file = f"../../config/{dbx_env}/{config_file_name}"
#config_file ="/Workspace/Users/Pawanpreet.Sangari@USPTO.GOV/bdr-trm-reports-dpl-jul25/notebooks/config/dev/trmreports-conf.yaml"
print(f"{config_file=}, {dbx_env=}")

# COMMAND ----------

# MAGIC %run ./../shared/ntb_common_func_and_params

# COMMAND ----------

common_configs = read_yaml(config_file)
reporting_catalog = common_configs["schema"]["trgt_catalog"]
tmngpdb_catalog = common_configs["schema"]["tmngpdb_src_catalog"]
tmngidmp_catalog = common_configs["schema"]["tmngidmp_catalog"]
trm_scope = common_configs["secrets"]["trm_scope"]
dq_catalog = common_configs['schema']['data_quality_catalog']
altrx_schema = common_configs['schema']['altrx_schema']
primary_email = common_configs["alerting"]["first_action_report_v1"]["email"]
edw_scope = common_configs['secrets']['edw_scope']
print(reporting_catalog, tmngpdb_catalog, tmngidmp_catalog, trm_scope, dq_catalog, altrx_schema, primary_email)

# COMMAND ----------

# set current time for job control
curntdt = datetime.datetime.now().astimezone(pytz.timezone('US/Eastern'))
run_date = curntdt.strftime('%Y-%m-%d')
# start job control  
job_start_ts = curntdt.strftime('%Y-%m-%d %H:%M:%S')
job_name = "ntb_tm_first_action_report_v1"
control_dt = begin_job_cntl(f"{reporting_catalog}.silver", job_name, job_start_ts)

# COMMAND ----------

input_edw_df = f"""SELECT DISTINCT
ACCTG_DT,
FEE_AM,
FEE_CD,
TRAN_PSTNG_REF_TX
FROM DW.VW_FPNG_SALE
WHERE
( LPAD(TRAN_PSTNG_REF_TX,1)='7'
OR   LPAD(TRAN_PSTNG_REF_TX,1)='8'
OR   LPAD(TRAN_PSTNG_REF_TX,1)='9'
OR   LPAD(TRAN_PSTNG_REF_TX,1)='A')
AND LENGTH(TRAN_PSTNG_REF_TX) = 8
AND FEE_CD IN ('7017','7018','7019','7020')
 """

new_fees_edw_df = read_data_from_oracle_conn_dsu_cmn(input_edw_df,edw_scope)
#new_fees_edw_df.display()

# COMMAND ----------

# Load the data from the 'pendency_dashboard' table for the specified fiscal year
df = spark.sql(f"SELECT * FROM {reporting_catalog}.gold.pendency_dashboard WHERE fa_pendency_fy = '{fiscal_year}'")

#Surcharge
df = df.join(
            new_fees_edw_df,
            on = [
                col("SER_NUM") == col("TRAN_PSTNG_REF_TX")
            ],
            how = "left"
        ).dropDuplicates()

# Add a new column 'Month_Sort' to the DataFrame based on the 'fa_pendency_fy_month' column
df = (
    df.withColumn("Month_Sort",  # New column 'Month_Sort'
                  # Use the 'when' function to assign numeric values to each month
                  when(col("fa_pendency_fy_month") == "Oct", 1)  # October -> 1
                  .when(col("fa_pendency_fy_month") == "Nov", 2)  # November -> 2
                  .when(col("fa_pendency_fy_month") == "Dec", 3)  # December -> 3
                  .when(col("fa_pendency_fy_month") == "Jan", 4)  # January -> 4
                  .when(col("fa_pendency_fy_month") == "Feb", 5)  # February -> 5
                  .when(col("fa_pendency_fy_month") == "Mar", 6)  # March -> 6
                  .when(col("fa_pendency_fy_month") == "Apr", 7)  # April -> 7
                  .when(col("fa_pendency_fy_month") == "May", 8)  # May -> 8
                  .when(col("fa_pendency_fy_month") == "Jun", 9)  # June -> 9
                  .when(col("fa_pendency_fy_month") == "Jul", 10)  # July -> 10
                  .when(col("fa_pendency_fy_month") == "Aug", 11)  # August -> 11
                  .when(col("fa_pendency_fy_month") == "Sep", 12)  # September -> 12
                  # If the month is not recognized, assign 0 as the default value
                  .otherwise(0)
                  )
)

# At this point, the DataFrame 'df' has a new column 'Month_Sort' which contains numeric values corresponding to the month names.
#df.display()

# COMMAND ----------

# -------------------- WORKFLOW 1 FIRST ACTION CLASSES --------------------

# Step 1: Grouping and Summing Active_Classes_FirstAction
# Group the DataFrame by fiscal year, month, and filing method, and sum the 'active_classes_firstaction' for each group
from pyspark.sql import Window
from pyspark.sql.functions import col, when, sum as _sum

df1 = df.groupBy(
    "fa_pendency_fy",  # Fiscal year
    "fa_pendency_fy_month",  # Fiscal year month
    "Month_Sort",  # Month sort order (numeric)
    "filing_method_filed"  # Filing method used
).agg(
    _sum("active_classes_firstaction").alias("Sum_Active_Classes_FirstAction"),  # Sum of active classes for first action
    _sum(when(col("fee_cd") == "7017", col("active_classes_firstaction")).otherwise(0)).alias("Sum_base"),  # Sum of base when fee_cd is 7017
    _sum(when(col("fee_cd").isin("7018", "7019", "7020"), col("active_classes_firstaction")).otherwise(0)).alias("Sum_base_surcharge")  # Sum of base_surcharge for fee_cd 7018, 7019, 7020
)
# Define a window specification for partitioning by 'fa_pendency_fy_month'
#window_spec = Window.partitionBy("fa_pendency_fy_month")

# Add columns to df1
df1 = (
    df1.withColumn("Paper", when(col("filing_method_filed") == "Paper", col("Sum_Active_Classes_FirstAction")))
        .withColumn("MADRID", when(col("filing_method_filed") == "MADRID", col("Sum_Active_Classes_FirstAction")))
        .withColumn("TEAS_PLUS", when(col("filing_method_filed") == "TEAS PLUS", col("Sum_Active_Classes_FirstAction")))
        .withColumn("TEAS_STD", when(col("filing_method_filed") == "TEAS STD", col("Sum_Active_Classes_FirstAction")))
        .withColumn("BASE", when(col("filing_method_filed") == "BASE", col("Sum_base")))
        .withColumn("BASE_SRCHRG", when(col("filing_method_filed") == "BASE", col("Sum_base_surcharge")))
)

# Coalesce Nulls to zeros
df1 = (
    df1.withColumn("Paper", coalesce(col("Paper"), lit(0)))
    .withColumn("MADRID", coalesce(col("MADRID"), lit(0)))
    .withColumn("TEAS_PLUS", coalesce(col("TEAS_PLUS"), lit(0)))
    .withColumn("TEAS_STD", coalesce(col("TEAS_STD"), lit(0)))
    .withColumn("BASE", coalesce(col("BASE"), lit(0)))
    .withColumn("BASE_SRCHRG", coalesce(col("BASE_SRCHRG"), lit(0)))
)

df1 = (
    df1.groupBy("fa_pendency_fy", "fa_pendency_fy_month", "Month_Sort")  # Grouping by fiscal year, month, and month sort order
    .agg(
        _sum("Paper").alias("Paper"),  # Sum of Paper filing method
        _sum("MADRID").alias("MADRID"),  # Sum of MADRID filing method
        _sum("TEAS_PLUS").alias("TEAS_PLUS"),  # Sum of TEAS PLUS filing method
        _sum("TEAS_STD").alias("TEAS_STD"),  # Sum of TEAS STD filing method
        _sum("BASE_SRCHRG").alias("BASE_SRCHRG"),
        _sum("BASE").alias("BASE")
    )
)

# Step 5: Sorting by Month_Sort
# Order the DataFrame by 'Month_Sort' to ensure the data is sorted by month
df1 = df1.orderBy("Month_Sort")

#display(df1)


# COMMAND ----------

# Step 6: Create Running Totals
# Define a window specification to order the data by 'Month_Sort' for calculating running totals
window_spec = Window.orderBy("Month_Sort")

# Calculate running totals for each filing method column (Paper, MADRID, TEAS_PLUS, TEAS_STD)
# The _sum function is used with the 'over' method to apply the window specification and compute cumulative sums
df1 = df1.withColumn("RunTot_Paper", _sum("Paper").over(window_spec)) \
    .withColumn("RunTot_MADRID", _sum("MADRID").over(window_spec)) \
    .withColumn("RunTot_TEAS_PLUS", _sum("TEAS_PLUS").over(window_spec)) \
    .withColumn("RunTot_TEAS_STD", _sum("TEAS_STD").over(window_spec)) \
    .withColumn("RunTot_BASE_SRCHRG", _sum("BASE_SRCHRG").over(window_spec)) \
    .withColumn("RunTot_BASE", _sum("BASE").over(window_spec))

# Step7: Add MMM-YY, Group, and Group_Total Columns
# Add a new column "MMM-YY" which combines the fiscal month and the last two digits of the fiscal year
# Using concat_ws to concatenate 'FA_Pendency_FY_Month' and the last two characters of 'FA_Pendency_FY' with a hyphen
df1 = df1.withColumn("MMM-YY", concat_ws("-", col("FA_Pendency_FY_Month"), col("FA_Pendency_FY").substr(-2, 2))) \
    .withColumn("Group", lit("FIRST ACTION CLASSES")) \
    .withColumn("Group_Total", col("RunTot_Paper") + col("RunTot_MADRID") + col("RunTot_TEAS_PLUS") + col("RunTot_TEAS_STD") + col("RunTot_BASE_SRCHRG") + col("RunTot_BASE"))

# Step 8: Rename and Select Final Columns
# Select the final columns to prepare the DataFrame for reporting or output
# Rename columns to provide more readable names and remove unnecessary ones
final_df1 = df1.select(
    col("MMM-YY").alias("Month"),  # Rename 'MMM-YY' to 'Month'
    col("RunTot_Paper").alias("Paper"),  # Rename 'RunTot_Paper' to 'Paper'
    col("RunTot_MADRID").alias("REP"),  # Rename 'RunTot_MADRID' to 'REP'
    col("RunTot_TEAS_STD").alias("TEAS"),  # Rename 'RunTot_TEAS_STD' to 'TEAS'
    col("RunTot_TEAS_PLUS").alias("TEAS Plus"),  # Rename 'RunTot_TEAS_PLUS' to 'TEAS Plus'
    col("RunTot_BASE_SRCHRG").alias("BASE SRCHRG"),
    col("RunTot_BASE").alias("BASE"),
    col("Group"),  # Keep 'Group' column as is
    col("Group_Total").alias("Total"),  # Rename 'Group_Total' to 'Total'
    col("Month_Sort")  # Keep 'Month_Sort' column for sorting or further use
)

# The final DataFrame, 'final_df1', is now ready for reporting or export, containing:
# - Monthly totals by filing method (Paper, MADRID, TEAS_PLUS, TEAS_STD)
# - Cumulative running totals for each filing method
# - The group and total for all filing methods combined

# COMMAND ----------

#df1.display()

# COMMAND ----------

# -------------------- WORKFLOW 2 FIRST ACTION PUBS --------------------
# Step 1: Filter Rows where first_action_type_num Column contains "APPROVED FOR PUB"
# Filter the DataFrame to include only rows where the 'first_action_type_num' column contains the string "APPROVED FOR PUB"
df2 = df.filter(col("first_action_type_num").contains("APPROVED FOR PUB"))


# Step 2: Grouping and Summing Active_Classes_FirstAction
# Group the filtered data by fiscal year, fiscal month, month sort order, and filing method
# Calculate the sum of 'active_classes_firstaction' for each group
df2 = df2.groupBy(
    "fa_Pendency_fy", "fa_pendency_fy_month", "Month_Sort", "filing_method_filed"
).agg(_sum("active_classes_firstaction").alias("Sum_Active_Classes_FirstAction"),
      _sum(when(col("fee_cd") == "7017", col("active_classes_firstaction")).otherwise(0)).alias("Sum_base"),   
      _sum(when(col("fee_cd").isin("7018", "7019", "7020"), col("active_classes_firstaction")).otherwise(0)).alias("Sum_base_surcharge"))

# Step 3: Add Columns for Each Filing Method (Paper, MADRID, TEAS_PLUS, TEAS_STD)
# Add new columns for each filing method, setting the sum of 'active_classes_firstaction' to the appropriate column based on 'filing_method_filed'
df2 = (
        df2.withColumn("Paper", when(col("filing_method_filed")=="Paper",col("Sum_Active_Classes_FirstAction")))
            .withColumn("MADRID", when(col("filing_method_filed")=="MADRID",col("Sum_Active_Classes_FirstAction")))
            .withColumn("TEAS_PLUS", when(col("filing_method_filed")=="TEAS PLUS",col("Sum_Active_Classes_FirstAction")))
            .withColumn("TEAS_STD", when(col("filing_method_filed")=="TEAS STD",col("Sum_Active_Classes_FirstAction")))
            .withColumn("BASE", when(col("filing_method_filed") == "BASE", col("Sum_base")))
            .withColumn("BASE_SRCHRG", when(col("filing_method_filed") == "BASE", col("Sum_base_surcharge")))
)

# Step 4: Select Required Columns
# Select the necessary columns, replacing null values in the newly created filing method columns with 0
df2 = df2.select(
    "FA_Pendency_FY",  # Fiscal Year
    "FA_Pendency_FY_Month",  # Fiscal Month
    "Month_Sort",  # Month Sort order
    when(col("Paper").isNull(), 0).otherwise(col("Paper")).alias("Paper"),  # Replace null values in 'Paper' column with 0
    when(col("MADRID").isNull(), 0).otherwise(col("MADRID")).alias("MADRID"),  # Replace null values in 'MADRID' column with 0
    when(col("TEAS_PLUS").isNull(), 0).otherwise(col("TEAS_PLUS")).alias("TEAS_PLUS"),  # Replace null values in 'TEAS_PLUS' column with 0
    when(col("TEAS_STD").isNull(), 0).otherwise(col("TEAS_STD")).alias("TEAS_STD"),  # Replace null values in 'TEAS_STD' column with 0
    when(col("BASE_SRCHRG").isNull(), 0).otherwise(col("BASE_SRCHRG")).alias("BASE_SRCHRG"),  # TEAS STD column, replace null with 0
    when(col("BASE").isNull(), 0).otherwise(col("BASE")).alias("BASE")
)

# Step 5: Grouping by Key Columns and Summing Filing Methods
# After selecting the required columns, group by fiscal year, fiscal month, and month sort order again
# Sum the values for each filing method to get the total for each group
df2 = (
    df2.groupBy("fa_pendency_fy", "fa_pendency_fy_month", "Month_Sort")
    .agg(
        _sum("Paper").alias("Paper"),  # Sum 'Paper' values for each group
        _sum("MADRID").alias("MADRID"),  # Sum 'MADRID' values for each group
        _sum("TEAS_PLUS").alias("TEAS_PLUS"),  # Sum 'TEAS_PLUS' values for each group
        _sum("TEAS_STD").alias("TEAS_STD"),  # Sum 'TEAS_STD' values for each group
        _sum("BASE_SRCHRG").alias("BASE_SRCHRG"),
        _sum("BASE").alias("BASE")
    )
)

# Step 6: Sorting by Month_Sort
# Sort the data by 'Month_Sort' to ensure the data is in chronological order
df2 = df2.orderBy("Month_Sort")

# Step 7: Create Running Totals
# Use a window specification to calculate cumulative sums (running totals) for each filing method
# The running totals are calculated using the '_sum' function and the 'over' method, based on the 'Month_Sort' column
df2 = df2.withColumn("RunTot_Paper", _sum("Paper").over(window_spec)) \
    .withColumn("RunTot_MADRID", _sum("MADRID").over(window_spec)) \
    .withColumn("RunTot_TEAS_PLUS", _sum("TEAS_PLUS").over(window_spec)) \
    .withColumn("RunTot_TEAS_STD", _sum("TEAS_STD").over(window_spec)) \
    .withColumn("RunTot_BASE_SRCHRG", _sum("BASE_SRCHRG").over(window_spec)) \
    .withColumn("RunTot_BASE", _sum("BASE").over(window_spec))

# Step 8: Add MMM-YY, Group, and Group_Total Columns
# Create the 'MMM-YY' column by concatenating the fiscal month and the last two digits of the fiscal year
# Add a static 'Group' column with the value "FIRST ACTION PUBS"
# Calculate the 'Total' by summing the running totals of all filing methods (Paper, MADRID, TEAS_PLUS, TEAS_STD)
df2 = df2.withColumn(
    "MMM-YY", concat_ws("-", col("FA_Pendency_FY_Month"), col("FA_Pendency_FY").substr(-2, 2))
).withColumn("Group", lit("FIRST ACTION PUBS")) \
    .withColumn("Total", col("RunTot_Paper") + col("RunTot_MADRID") + col("RunTot_TEAS_PLUS") + col("RunTot_TEAS_STD")+ col("RunTot_BASE_SRCHRG") +col("RunTot_BASE"))

# Step 9: Rename and Select Final Columns
# Select the final columns for reporting, renaming them to provide clearer names for each filing method and the total
# Also, include the 'Month_Sort' column for sorting purposes
final_df2 = df2.select(
    col("MMM-YY").alias("Month"),  # Rename 'MMM-YY' to 'Month'
    col("RunTot_Paper").alias("Paper"),  # Rename 'RunTot_Paper' to 'Paper'
    col("RunTot_MADRID").alias("REP"),  # Rename 'RunTot_MADRID' to 'REP' (for representation)
    col("RunTot_TEAS_STD").alias("TEAS"),  # Rename 'RunTot_TEAS_STD' to 'TEAS'
    col("RunTot_TEAS_PLUS").alias("TEAS Plus"),  # Rename 'RunTot_TEAS_PLUS' to 'TEAS Plus'
    col("RunTot_BASE_SRCHRG").alias("BASE SRCHRG"),
    col("RunTot_BASE").alias("BASE"),
    col("Group"),  # Keep 'Group' column as is
    col("Total").alias("Total"),  # Rename 'Total' column to 'Total'
    col("Month_Sort")  # Keep 'Month_Sort' column for sorting or further use
)

# The final DataFrame, 'final_df2', is now ready for reporting or export, containing:
# - Monthly totals by filing method (Paper, MADRID, TEAS_PLUS, TEAS_STD)
# - Cumulative running totals for each filing method
# - The group and total for all filing methods combined

# COMMAND ----------

#df2.display()

# COMMAND ----------

# -------------------- WORKFLOW 3 FIRST ACTION PUB PERCENTAGES --------------------
# Step 1: Join df1 and df2 on Month_Sort to align data for percentage calculations
# This ensures that the data from both datasets (df1 and df2) are matched by the "Month" column for further calculations
joined_df = final_df1.alias("df1").join(final_df2.alias("df2"), on="Month", how="inner")

# Step 2: Calculate percentages for each category (Paper, REP, TEAS, TEAS Plus, Total)
# For each column, we divide the corresponding value in df2 by the value in df1, multiply by 100, 
# round the result to 1 decimal place, and concatenate the percentage symbol "%"
final_df3 = joined_df.select(
    col("df2.Month").alias("Month"),  # Select Month from df2 (as both df1 and df2 have this column)
    
    # Calculate and format percentage for Paper
    concat(round((col("df2.Paper") / col("df1.Paper")) * 100, 2), lit("%")).alias("Paper"),
    
    # Calculate and format percentage for REP
    concat(round((col("df2.REP") / col("df1.REP")) * 100, 2), lit("%")).alias("REP"),
    
    # Calculate and format percentage for TEAS
    concat(round((col("df2.TEAS") / col("df1.TEAS")) * 100, 2), lit("%")).alias("TEAS"),
    
    # Calculate and format percentage for TEAS Plus
    concat(round((col("df2.TEAS Plus") / col("df1.TEAS Plus")) * 100, 2), lit("%")).alias("TEAS Plus"),
    
    # Calculate and format percentage for Base
    concat(round((col("df2.BASE") / col("df1.BASE")) * 100, 2), lit("%")).alias("BASE"),
    
     # Calculate and format percentage for Base Surcharge
    concat(round((col("df2.BASE SRCHRG") / col("df1.BASE SRCHRG")) * 100, 2), lit("%")).alias("BASE Surcharge"),

    # Include Group column from df2 (no calculation, just inclusion)
    col("df2.Group").alias("Group"),
    
    # Calculate and format percentage for Total (sum of all categories)
    concat(round((col("df2.Total") / col("df1.Total")) * 100, 2), lit("%")).alias("Total"),
    
    # Include Month_Sort column from df2 (no calculation, just inclusion)
    col("df2.Month_Sort").alias("Month_Sort")
)


# COMMAND ----------

#final_df3.display()

# COMMAND ----------

# -------------------- COMBINE DATA FROM df1, df2, and df3 --------------------
# Step 1: Join df1, df2, and df3 on the "Month" column
# The join type is "inner," meaning only rows with matching months across all three DataFrames will be included in the result
combined_df = final_df1.alias("df1") \
    .join(final_df2.alias("df2"), on="Month", how="inner") \
    .join(final_df3.alias("df3"), on="Month", how="inner")

# Step 2: Select and format the columns from df1, df2, and df3
# Each selected column is formatted using the `format_number` function to round the numbers to 0 decimal places
# The format_number function ensures that numeric columns are displayed without decimal places in the final result
final_combined_df = combined_df.select(
    col("df1.Month").alias("Month"),  # Month column from df1 (same in all DataFrames)
    
    # Format and select the Paper, REP, TEAS, TEAS Plus, and Total columns from df1
    format_number(col("df1.Paper"), 0).alias("Paper"),
    format_number(col("df1.REP"), 0).alias("REP"),
    format_number(col("df1.TEAS"), 0).alias("TEAS"),
    format_number(col("df1.TEAS Plus"), 0).alias("TEAS Plus"),
    format_number(col("df1.BASE"), 0).alias("BASE"),
    format_number(col("df1.BASE SRCHRG"), 0).alias("BASE Surcharge"),
    format_number(col("df1.Total"), 0).alias("Total"),
    
    # Format and select the Paper, REP, TEAS, TEAS Plus, and Total columns from df2
    format_number(col("df2.Paper"), 0).alias("Paper"),
    format_number(col("df2.REP"), 0).alias("REP"),
    format_number(col("df2.TEAS"), 0).alias("TEAS"),
    format_number(col("df2.TEAS Plus"), 0).alias("TEAS Plus"),
    format_number(col("df2.BASE"), 0).alias("BASE"),
    format_number(col("df2.BASE SRCHRG"), 0).alias("BASE Surcharge"),
    format_number(col("df2.Total"), 0).alias("Total"),
    
    # Select the Paper, REP, TEAS, TEAS Plus, and Total columns from df3 without formatting
    col("df3.Paper").alias("Paper"),
    col("df3.REP").alias("REP"),
    col("df3.TEAS").alias("TEAS"),
    col("df3.TEAS Plus").alias("TEAS Plus"),
    col("df3.BASE").alias("BASE"),
    col("df3.BASE Surcharge").alias("BASE Surcharge"),
    col("df3.Total").alias("Total")
)


# COMMAND ----------

#final_combined_df.display()

# COMMAND ----------

import builtins
from openpyxl.utils import get_column_letter

pd_df = final_combined_df.toPandas()

if pd_df.empty:
    raise ValueError("The DataFrame is empty. Please provide valid data.")

wb = Workbook()
ws = wb.active
ws.title = "TM First Action Report"

header_fill = PatternFill(start_color="003865", end_color="003865", fill_type="solid")
header_font = Font(name="Arial", bold=True, size=14, color="004C97")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
)
bold_border = Border(
    left=Side(style="medium"), right=Side(style="medium"), top=Side(style="medium"), bottom=Side(style="medium")
)
data_font = Font(name="Arial")
footer_font = Font(name="Arial", italic=True)

ws.row_dimensions[1].height = 75

image_left_path = "../shared/uspto_logo.png"
image_right_path = "../shared/tm_analytics.jpg"

img_left = Image(image_left_path)
img_right = Image(image_right_path)
img_left.width, img_left.height = 130, 43 
img_right.width, img_right.height = 70, 90 

# Dynamically determine last column letter
num_columns = len(pd_df.columns)
last_col_letter = get_column_letter(num_columns + 1)  # +1 because you start at column 2 (B)

ws.add_image(img_left, "B1")
ws.add_image(img_right, f"{last_col_letter}1")

# Title for the report (Row 1)
ws.merge_cells(f"C1:{get_column_letter(num_columns)}1")
ws["C1"] = "TM FIRST ACTION (INITIAL EXAM) PUBLICATION REPORT"
ws["C1"].font = Font(bold=True, size=18, color="FFFFFF")
ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
for cell in ws[f"C1:{get_column_letter(num_columns)}1"][0]:
    cell.fill = header_fill
for cell in ws[f"B1:{last_col_letter}1"][0]:
    cell.border = bold_border

# Subtitles for the sections (Row 2) - explicit column ranges
# FIRST ACTION CLASSES: B2 to I2
ws.merge_cells("B2:I2")
ws["B2"] = "FIRST ACTION CLASSES"
ws["B2"].font = header_font
ws["B2"].alignment = Alignment(horizontal="center")
for cell in ws["B2:I2"][0]:
    cell.border = bold_border

# FIRST ACTION INITIAL PUBS: J2 to P2
ws.merge_cells("J2:P2")
ws["J2"] = "FIRST ACTION INITIAL PUBS"
ws["J2"].font = header_font
ws["J2"].alignment = Alignment(horizontal="center")
for cell in ws["J2:P2"][0]:
    cell.border = bold_border

# FYTD FIRST ACTION PUB PERCENTAGE: Q2 to W2
ws.merge_cells("Q2:W2")
ws["Q2"] = "FYTD FIRST ACTION PUB PERCENTAGE"
ws["Q2"].font = header_font
ws["Q2"].alignment = Alignment(horizontal="center")
for cell in ws["Q2:W2"][0]:
    cell.border = bold_border

# Add headers (Row 3)
for col_num, header in enumerate(pd_df.columns, 1):
    cell = ws.cell(row=3, column=col_num + 1, value=header)
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center" if header == "Month" else "left")
    cell.border = bold_border

# Add data rows starting from Row 4
for row_num, row_data in enumerate(pd_df.itertuples(index=False), start=4):
    for col_num, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_num, column=col_num + 1, value=value)
        cell.font = data_font
        if pd_df.columns[col_num - 1] == "Month":
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.alignment = Alignment(horizontal="left")
        cell.border = thin_border

# Adjust column widths
for col_num, column_cells in enumerate(ws.iter_cols(min_row=3, max_row=ws.max_row), start=1):
    try:
        max_length = builtins.max(
            len(str(cell.value)) for cell in column_cells if cell.value
        ) + 2
    except ValueError:
        max_length = 2
    ws.column_dimensions[get_column_letter(col_num + 1)].width = builtins.max(max_length, 18)

# Footer
footer_row = ws.max_row + 2
ws.merge_cells(f"B{footer_row}:{last_col_letter}{footer_row}")
ws[f"B{footer_row}"] = f"Run date: {pd.Timestamp.now().strftime('%Y-%m-%d')} - Report generated from Databricks - TM First Action Report"
ws[f"B{footer_row}"].alignment = Alignment(horizontal="left")
ws[f"B{footer_row}"].font = footer_font

# Hyperlink
hyperlink_row = footer_row + 2
ws.merge_cells(f"B{hyperlink_row}:{last_col_letter}{hyperlink_row}")
hyperlink_cell = ws[f"B{hyperlink_row}"]
hyperlink_cell.value = "Click to submit a TM DnA Request"
hyperlink_cell.font = Font(name="Arial", underline="single", color="0000FF")
hyperlink_cell.alignment = Alignment(horizontal="left")
hyperlink_cell.hyperlink = "https://apps.gov.powerapps.us/play/e/default-ff4abfe9-83b5-4026-8b8f-fa69a1cad0b8/a/ea18fb4c-aa64-4056-b0d5-f5ad4097cf0d?tenantId=ff4abfe9-83b5-4026-8b8f-fa69a1cad0b8&source=email&sourcetime=1731514719403"

report_name = "EP_115"
local_file_output_path = f"/tmp/{report_name}.xlsx"
dbfs_file_output_path = f"/dbfs/mnt/eds/trademark/dbx_reports/first_action_publication/{report_name}.xlsx"

try:
    wb.save(local_file_output_path)
    shutil.move(local_file_output_path, dbfs_file_output_path)
    print(f"File saved to {dbfs_file_output_path}")
except Exception as e:
    print(f"Error while saving or moving the file: {e}")

# COMMAND ----------

import builtins
from openpyxl.utils import get_column_letter

pd_df = final_combined_df.toPandas()

if pd_df.empty:
    raise ValueError("The DataFrame is empty. Please provide valid data.")

wb = Workbook()
ws = wb.active
ws.title = "TM First Action Report"

header_fill = PatternFill(start_color="003865", end_color="003865", fill_type="solid")
header_font = Font(name="Arial", bold=True, size=14, color="004C97")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
)
bold_border = Border(
    left=Side(style="medium"), right=Side(style="medium"), top=Side(style="medium"), bottom=Side(style="medium")
)
data_font = Font(name="Arial")
footer_font = Font(name="Arial", italic=True)

ws.row_dimensions[1].height = 75

image_left_path = "../shared/uspto_logo.png"
image_right_path = "../shared/tm_analytics.jpg"

img_left = Image(image_left_path)
img_right = Image(image_right_path)
img_left.width, img_left.height = 130, 43 
img_right.width, img_right.height = 70, 90 

# Dynamically determine last column letter
num_columns = len(pd_df.columns)
last_col_letter = get_column_letter(num_columns + 1)  # +1 because you start at column 2 (B)

ws.add_image(img_left, "B1")
ws.add_image(img_right, f"{last_col_letter}1")

# Title for the report (Row 1)
ws.merge_cells(f"C1:{get_column_letter(num_columns)}1")
ws["C1"] = "TM FIRST ACTION (INITIAL EXAM) PUBLICATION REPORT"
ws["C1"].font = Font(bold=True, size=18, color="FFFFFF")
ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
for cell in ws[f"C1:{get_column_letter(num_columns)}1"][0]:
    cell.fill = header_fill
for cell in ws[f"B1:{last_col_letter}1"][0]:
    cell.border = bold_border

# Subtitles for the sections (Row 2) - explicit column ranges
# FIRST ACTION CLASSES: B2 to I2
ws.merge_cells("B2:I2")
ws["B2"] = "FIRST ACTION CLASSES"
ws["B2"].font = header_font
ws["B2"].alignment = Alignment(horizontal="center")
for cell in ws["B2:I2"][0]:
    cell.border = bold_border

# FIRST ACTION INITIAL PUBS: J2 to P2
ws.merge_cells("J2:P2")
ws["J2"] = "FIRST ACTION INITIAL PUBS"
ws["J2"].font = header_font
ws["J2"].alignment = Alignment(horizontal="center")
for cell in ws["J2:P2"][0]:
    cell.border = bold_border

# FYTD FIRST ACTION PUB PERCENTAGE: Q2 to W2
ws.merge_cells("Q2:W2")
ws["Q2"] = "FYTD FIRST ACTION PUB PERCENTAGE"
ws["Q2"].font = header_font
ws["Q2"].alignment = Alignment(horizontal="center")
for cell in ws["Q2:W2"][0]:
    cell.border = bold_border

# Add headers (Row 3)
for col_num, header in enumerate(pd_df.columns, 1):
    cell = ws.cell(row=3, column=col_num + 1, value=header)
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center" if header == "Month" else "left")
    cell.border = bold_border

# Add data rows starting from Row 4
for row_num, row_data in enumerate(pd_df.itertuples(index=False), start=4):
    for col_num, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_num, column=col_num + 1, value=value)
        cell.font = data_font
        if pd_df.columns[col_num - 1] == "Month":
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.alignment = Alignment(horizontal="left")
        cell.border = thin_border

# Adjust column widths
for col_num, column_cells in enumerate(ws.iter_cols(min_row=3, max_row=ws.max_row), start=1):
    try:
        max_length = builtins.max(
            len(str(cell.value)) for cell in column_cells if cell.value
        ) + 2
    except ValueError:
        max_length = 2
    ws.column_dimensions[get_column_letter(col_num + 1)].width = builtins.max(max_length, 18)

# Footer
footer_row = ws.max_row + 2
ws.merge_cells(f"B{footer_row}:{last_col_letter}{footer_row}")
ws[f"B{footer_row}"] = f"Run date: {pd.Timestamp.now().strftime('%Y-%m-%d')} - Report generated from Databricks - TM First Action Report"
ws[f"B{footer_row}"].alignment = Alignment(horizontal="left")
ws[f"B{footer_row}"].font = footer_font

# Hyperlink
hyperlink_row = footer_row + 2
ws.merge_cells(f"B{hyperlink_row}:{last_col_letter}{hyperlink_row}")
hyperlink_cell = ws[f"B{hyperlink_row}"]
hyperlink_cell.value = "Click to submit a TM DnA Request"
hyperlink_cell.font = Font(name="Arial", underline="single", color="0000FF")
hyperlink_cell.alignment = Alignment(horizontal="left")
hyperlink_cell.hyperlink = "https://apps.gov.powerapps.us/play/e/default-ff4abfe9-83b5-4026-8b8f-fa69a1cad0b8/a/ea18fb4c-aa64-4056-b0d5-f5ad4097cf0d?tenantId=ff4abfe9-83b5-4026-8b8f-fa69a1cad0b8&source=email&sourcetime=1731514719403"

report_name = "EP_115"
local_file_output_path = f"/tmp/{report_name}.xlsx"
dbfs_file_output_path = f"/dbfs/mnt/eds/trademark/dbx_reports/first_action_publication/{report_name}.xlsx"

try:
    wb.save(local_file_output_path)
    shutil.move(local_file_output_path, dbfs_file_output_path)
    print(f"File saved to {dbfs_file_output_path}")
except Exception as e:
    print(f"Error while saving or moving the file: {e}")

# COMMAND ----------

# Email credentials
to = primary_email
from_addr = "trademark_analytics@uspto.gov"
subj = f"First Action Publication Report: {run_date}"
html = f"""
        Please find attached the First Action Publication Report in Excel format.<br><br>
        Best Regards,<br><br>
        Trademark DnA Team
        """

notify = Notify()

# Attach the PDF file
attachments = [dbfs_file_output_path]  

# Compose the email with the attachment
msg = notify.compose_email_attachment_with_html_body(
    html=html,
    subj=subj,
    to=to,
    from_addr=from_addr,
    filepaths=attachments  # Attach the file using the file path
)

# Send the email
notify.send_mail(msg)

# COMMAND ----------

table_combined_df = combined_df.select(
    col("df1.Month").alias("Month"),  # Month column from df1 (same in all DataFrames)

    # Format and select the Paper, REP, TEAS, TEAS Plus, and Total columns from df1
    format_number(col("df1.Paper"), 0).alias("Paper_classes"),
    format_number(col("df1.REP"), 0).alias("REP_classes"),
    format_number(col("df1.TEAS"), 0).alias("TEAS_classes"),
    format_number(col("df1.TEAS Plus"), 0).alias("TEAS_Plus_classes"),
    format_number(col("df1.Total"), 0).alias("Total_classes"),
    
    # Format and select the Paper, REP, TEAS, TEAS Plus, and Total columns from df2
    format_number(col("df2.Paper"), 0).alias("Paper_pubs"),
    format_number(col("df2.REP"), 0).alias("REP_pubs"),
    format_number(col("df2.TEAS"), 0).alias("TEAS_pubs"),
    format_number(col("df2.TEAS Plus"), 0).alias("TEAS_Plus_pubs"),
    format_number(col("df2.Total"), 0).alias("Total_pubs"),
    
    # Select the Paper, REP, TEAS, TEAS Plus, and Total columns from df3 without formatting
    col("df3.Paper").alias("Paper_pub_percent"),
    col("df3.REP").alias("REP_pub_percent"),
    col("df3.TEAS").alias("TEAS_pub_percent"),
    col("df3.TEAS Plus").alias("TEAS_Plus_pub_percent"),
    col("df3.Total").alias("Total_pub_percent")
)

table_combined_df.write.mode("overwrite").format("delta").saveAsTable(f"{reporting_catalog}.gold.first_action_report")

# COMMAND ----------

#display(table_combined_df)

# COMMAND ----------

end_job_cntl(
    f"{reporting_catalog}.silver",
    job_name,
    job_start_ts,
    "completed",
    0,
    "job completed successfully",
)
dbutils.notebook.exit(f"Job completed with {df.count()} records.")
